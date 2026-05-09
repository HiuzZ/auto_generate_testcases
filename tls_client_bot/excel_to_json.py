#!/usr/bin/env python3
import argparse
import json
import re
from collections import OrderedDict

from openpyxl import load_workbook


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _split_numbered_list(text):
    """
    Split multiline numbered text like:
    1. item1\n2. item2
    Returns a list of items without the number prefix.
    """
    if not text:
        return []

    lines = str(text).splitlines()
    results = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # remove numbering like "1. ", "2) ", etc.
        line = re.sub(r"^\d+[\.\)]\s*", "", line)
        if line and line not in results:
            results.append(line)

    return results


def _split_responses(text):
    """Alias for splitting responses (same behaviour)."""
    return _split_numbered_list(text)


def _split_patterns(text):
    """
    If text contains newlines and looks like a numbered list,
    split it using _split_numbered_list.
    Otherwise, return as a single-element list if not empty.
    """
    if not text:
        return []

    # Kiểm tra có xuống dòng và dòng đầu tiên bắt đầu bằng số + dấu chấm/đóng ngoặc
    if "\n" in text and re.search(r"^\d+[\.\)]", text, re.MULTILINE):
        return _split_numbered_list(text)
    else:
        # Không phải danh sách đánh số, giữ nguyên một pattern
        return [text] if text else []


def parse_excel_to_intents(excel_path, sheet_name=None):
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    intents_map = OrderedDict()
    current_tag = None

    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        # Handle flexible number of columns
        intent_cell   = row[0] if len(row) > 0 else None
        pattern_cell  = row[1] if len(row) > 1 else None  # column B – patterns
        response_cell = row[2] if len(row) > 2 else None  # column C – responses

        tag      = _clean_text(intent_cell)
        # patterns: xử lý đặc biệt để hỗ trợ cả ô gộp nhiều pattern
        patterns = _split_patterns(_clean_text(pattern_cell))
        response = _clean_text(response_cell)

        if tag:
            current_tag = tag

        if not current_tag:
            continue

        if current_tag not in intents_map:
            intents_map[current_tag] = {
                "tag": current_tag,
                "patterns": [],
                "responses": [],
            }

        # Add each pattern from the (possibly multiple) list
        for pat in patterns:
            if pat not in intents_map[current_tag]["patterns"]:
                intents_map[current_tag]["patterns"].append(pat)

        # Add responses (column C) – split numbered lists
        for r in _split_responses(response):
            if r not in intents_map[current_tag]["responses"]:
                intents_map[current_tag]["responses"].append(r)

    return {"intents": list(intents_map.values())}


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel voice-flow data to intents JSON format."
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="Path to source Excel file.",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="intents_from_excel.json",
        help="Path to output JSON file (default: intents_from_excel.json).",
    )
    parser.add_argument(
        "--sheet",
        "-s",
        default=None,
        help="Sheet name (default: first sheet).",
    )
    args = parser.parse_args()

    intents_data = parse_excel_to_intents(args.input, args.sheet)

    with open(args.output, "w", encoding="utf-8") as output_file:
        json.dump(intents_data, output_file, ensure_ascii=False, indent=2)

    print(f"Created JSON with {len(intents_data['intents'])} intents: {args.output}")


if __name__ == "__main__":
    main()