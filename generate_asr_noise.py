"""
generate_asr_noise.py — Inject realistic Vietnamese ASR mistakes into generated test data.

Primary method: Llama 3.1 via Ollama (same as gen_data LLM/Hybrid modes).
Fallback: rule-based word/tone swaps if Ollama is unavailable.

Rules:
- Targets ONLY the "User Message" column. No other column is touched.
- Skips cells that are empty or equal to _SILENCE_.
- Plain-text cell (no numbered lines): rewrites the whole cell value.
- Numbered multi-line cell: picks exactly ONE random eligible step to rewrite.
"""
from __future__ import annotations

import random
import re
from pathlib import Path
from typing import Callable, Optional

import openpyxl as xl
from openpyxl import Workbook

try:
    from ollama import chat as ollama_chat
except ImportError:
    ollama_chat = None

LLAMA_MODEL = "llama3.1"

_ASR_NOISE_PROMPT = (
    "Chuyển câu nói của khách hàng sau thành phiên bản có lỗi nhận dạng giọng nói (ASR mistake) "
    "thực tế trong tiếng Việt. Giữ nguyên nghĩa, chỉ thay 1-2 từ bằng từ đồng âm hoặc gần âm "
    "mà hệ thống ASR thường nhầm (ví dụ: \"hà nội\" → \"hà lội\", \"vay\" → \"dây\", "
    "\"không\" → \"hông\", \"tôi\" → \"tui\", \"lãi suất\" → \"lái suất\").\n\n"
    "Câu gốc: {utterance}\n\n"
    "Chỉ trả về câu đã chuyển đổi, không giải thích, không thêm dấu ngoặc kép."
)

# ── Fallback rule-based tables ─────────────────────────────────────────────────

_WORD_SWAPS: list[tuple[str, str]] = [
    ("hà nội", "hà lội"),
    ("nội dung", "lội dung"),
    ("không", "hông"),
    ("được", "đươc"),
    ("muốn", "muống"),
    ("vay tiền", "dây tiền"),
    ("vay", "dây"),
    ("tiền", "tiên"),
    ("triệu", "trệu"),
    ("đồng", "dồng"),
    ("tỷ", "tý"),
    ("tỉ", "tí"),
    ("khoản", "khỏan"),
    ("lãi suất", "lái suất"),
    ("lãi", "lái"),
    ("tín dụng", "tin dụng"),
    ("thông tin", "thông tính"),
    ("ngân hàng", "ngân háng"),
    ("tài khoản", "tài khỏan"),
    ("gửi tiết kiệm", "gởi tiết kiệm"),
    ("chuyển khoản", "chuyển khỏan"),
    ("bảo hiểm", "báo hiểm"),
    ("rút tiền", "giút tiền"),
    ("nạp tiền", "nặp tiền"),
    ("thẻ tín dụng", "thẻ tin dụng"),
    ("đăng ký", "đăng ki"),
    ("hồ sơ", "hổ sơ"),
    ("thủ tục", "thú tục"),
    ("lương", "lươn"),
    ("trả góp", "trả gộp"),
    ("phí", "phỉ"),
    ("vốn", "vống"),
    ("tôi muốn", "tui muốn"),
    ("tôi cần", "tui cần"),
    ("tôi không", "tui không"),
    ("tôi đang", "tui đang"),
    ("tôi có", "tui có"),
    ("tôi ở", "tui ở"),
    ("mình", "minh"),
    ("năm triệu", "lăm triệu"),
    ("mười", "mươi"),
    ("bốn", "bổn"),
    ("một", "mốt"),
    ("làm", "nàm"),
    ("đang", "đăng"),
    ("cần", "căn"),
    ("tháng", "thám"),
    ("năm", "lăm"),
    ("như thế nào", "như thế lào"),
    ("bao nhiêu", "bao liêu"),
    ("hỗ trợ", "hổ trợ"),
    ("có thể", "cổ thể"),
    ("nhanh", "nanh"),
    ("biết", "biếc"),
    ("hiểu", "hiêu"),
    ("giúp", "dúp"),
    ("xử lý", "sử lý"),
    ("thông báo", "thông bão"),
    ("thanh toán", "than toán"),
]

_TONE_MAP: dict[str, str] = {
    "ộ": "ọ", "ọ": "ộ",
    "ổ": "ỗ", "ỗ": "ổ",
    "ề": "ế", "ế": "ề",
    "ắ": "ặ", "ặ": "ắ",
    "ớ": "ờ", "ờ": "ớ",
    "ứ": "ừ", "ừ": "ứ",
    "ị": "ỉ", "ỉ": "ị",
    "ẫ": "ẩ", "ẩ": "ẫ",
    "ẹ": "ẽ", "ẽ": "ẹ",
    "ụ": "ủ", "ủ": "ụ",
}

_STEP_RE = re.compile(r"^(\d+\.\s*)")


# ── Rule-based fallback ────────────────────────────────────────────────────────

def _apply_word_swap(text: str, rng: random.Random) -> str:
    text_lower = text.lower()
    candidates = [(orig, mis) for orig, mis in _WORD_SWAPS if orig in text_lower]
    if not candidates:
        return text
    orig, mistake = rng.choice(candidates)
    idx = text_lower.find(orig)
    return text[:idx] + mistake + text[idx + len(orig):]


def _apply_tone_swap(text: str, rng: random.Random) -> str:
    positions = [(i, c) for i, c in enumerate(text) if c in _TONE_MAP]
    if not positions:
        return text
    i, c = rng.choice(positions)
    return text[:i] + _TONE_MAP[c] + text[i + 1:]


def _rule_based_noise(utterance: str, rng: random.Random) -> str:
    noisy = _apply_word_swap(utterance, rng)
    if noisy != utterance:
        return noisy
    return _apply_tone_swap(utterance, rng)


# ── Ollama/LLaMA primary method ────────────────────────────────────────────────

def _llama_noise(utterance: str) -> Optional[str]:
    """Ask LLaMA to produce a realistic ASR-confused version. Returns None on failure."""
    if not utterance or not utterance.strip() or ollama_chat is None:
        return None
    prompt = _ASR_NOISE_PROMPT.format(utterance=utterance)
    try:
        response = ollama_chat(
            model=LLAMA_MODEL,
            messages=[{"role": "user", "content": prompt}],
        )
        content = (response.get("message") or {}).get("content") or ""
        content = content.strip().strip('"').strip("'").strip()
        if content and content.lower() != utterance.lower():
            return content
    except Exception as exc:
        print(f"    [WARN] Ollama ASR noise failed: {exc}")
    return None


def add_asr_noise(utterance: str, rng: Optional[random.Random] = None) -> str:
    """Return a version of *utterance* with a realistic Vietnamese ASR mistake.

    Uses LLaMA via Ollama when available; falls back to rule-based swaps.
    """
    if not utterance or not utterance.strip():
        return utterance
    if rng is None:
        rng = random.Random()
    result = _llama_noise(utterance)
    if result is not None:
        return result
    return _rule_based_noise(utterance, rng)


# ── Cell-level processing ──────────────────────────────────────────────────────

def _find_col_in_row(row: tuple, target: str) -> Optional[int]:
    """Return 1-based column index of the first cell whose text contains *target* (case-insensitive)."""
    for idx, val in enumerate(row, start=1):
        if val and target.lower() in str(val).lower():
            return idx
    return None


def _noise_user_message_cell(cell_value: str, rng: random.Random) -> tuple[str, int]:
    """Apply ASR noise to a single User Message cell.

    - Empty cell or _SILENCE_: skipped, returned unchanged.
    - Plain text (no numbered prefix): the whole value is one utterance → rewrite it.
    - Numbered multi-line ("1. ...\n2. ..."): pick exactly ONE eligible step at random.

    Returns (new_value, 1) if the cell was changed, (original, 0) otherwise.
    """
    if not cell_value or not cell_value.strip():
        return cell_value, 0

    stripped = cell_value.strip()
    if stripped.upper() == "_SILENCE_":
        return cell_value, 0

    lines = cell_value.split("\n")

    # Track all numbered steps and separately the eligible (non-SILENCE) ones
    numbered: list[int] = []
    eligible: list[int] = []
    for i, line in enumerate(lines):
        m = _STEP_RE.match(line)
        if not m:
            continue
        numbered.append(i)
        utterance = line[len(m.group(1)):]
        if utterance.strip() and utterance.strip().upper() != "_SILENCE_":
            eligible.append(i)

    if numbered:
        # Numbered format — skip entire cell if every step is _SILENCE_
        if not eligible:
            return cell_value, 0
        chosen = rng.choice(eligible)
        m = _STEP_RE.match(lines[chosen])
        assert m
        prefix = m.group(1)
        utterance = lines[chosen][len(prefix):]
        noisy = add_asr_noise(utterance, rng)
        if noisy == utterance:
            return cell_value, 0
        lines[chosen] = prefix + noisy
        return "\n".join(lines), 1
    else:
        # Plain text — treat the whole cell as one utterance
        noisy = add_asr_noise(stripped, rng)
        if noisy == stripped:
            return cell_value, 0
        return noisy, 1


# ── Excel entry point ──────────────────────────────────────────────────────────

def fill_excel_with_asr_noise(
    input_path: Path,
    output_path: Path,
    progress_fn: Optional[Callable[[int, int], None]] = None,
) -> int:
    """Read *input_path*, apply ASR noise to the "User Message" column, write *output_path*.

    Searches the first 6 rows for a header containing "user message" to handle
    multi-row / merged-cell Excel layouts. Only that column is modified.

    *progress_fn(done, total)* is called after each data row when provided.

    Returns the number of rows where the cell was rewritten.
    """
    wb: Workbook = xl.load_workbook(input_path)
    modified_total = 0

    for ws in wb.worksheets:
        # Search header rows (up to row 6) for "User Message"
        target_col: Optional[int] = None
        header_row_idx = 1
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1
        ):
            col = _find_col_in_row(row, "user message")  # type: ignore[arg-type]
            if col is not None:
                target_col = col
                header_row_idx = row_idx
                break

        if target_col is None:
            continue  # no "User Message" column in this sheet — skip entirely

        rng = random.Random(42)
        data_rows = list(ws.iter_rows(min_row=header_row_idx + 1))
        total = len(data_rows)

        for done, row in enumerate(data_rows, start=1):
            cell = row[target_col - 1]
            original = cell.value
            if original and str(original).strip():
                noisy, count = _noise_user_message_cell(str(original), rng)
                if count > 0:
                    cell.value = noisy
                    modified_total += count
            if progress_fn is not None and total > 0:
                progress_fn(done, total)

    wb.save(output_path)
    return modified_total
