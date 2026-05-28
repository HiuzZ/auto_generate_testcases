from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def workbook_to_json(
    path: Path,
    *,
    sheet: str | int | None = None,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_names = wb.sheetnames
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
    else:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_str(c) for c in row])

    return {
        "path": str(path.name),
        "sheet_names": sheet_names,
        "active_sheet": ws.title,
        "rows": rows,
        "row_count": len(rows),
        "col_count": max((len(r) for r in rows), default=0),
    }


def workbook_from_json(
    *,
    sheet_names: list[str],
    active_sheet: str,
    rows: list[list[str]],
    dest: Path,
) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    title = active_sheet if active_sheet in sheet_names else (sheet_names[0] if sheet_names else "Sheet1")
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
    for name in sheet_names:
        if name != title and name not in wb.sheetnames:
            wb.create_sheet(title=name)
    wb.save(dest)
    return dest
