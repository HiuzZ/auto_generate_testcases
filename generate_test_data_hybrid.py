"""
Generate realistic customer conversation utterances for each test case using a hybrid approach:
1. First try TLS Client Bot (NLP intent classifier)
2. If TLS bot fails for a specific step, fallback to Llama 3.1 (Ollama) for that step only

Bot responses may contain multiple variants separated by " \ "; one is chosen randomly per generation.
"""

from __future__ import annotations

from collections.abc import Iterable
import argparse
import json
import os
import random
import re
import sys
from contextlib import contextmanager
from pathlib import Path
from typing import Any
import unicodedata

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Try to import Ollama (optional dependency)
try:
    from ollama import chat as ollama_chat
except ImportError:
    ollama_chat = None

# TLS Client Bot imports
DEFAULT_MODEL = "tls_client_bot"
MAX_WORDS_PER_RESPONSE = 50

PROJECT_ROOT = Path(__file__).resolve().parent
TLS_BOT_DIR = PROJECT_ROOT / "tls_client_bot"
TLS_INTENTS_PATH = TLS_BOT_DIR / "intents.json"

_TLS_GET_RESPONSE = None
_TLS_INTENTS_DATA: dict[str, Any] | None = None
_TAGS_SET: set[str] | None = None
_PATTERN_TO_TAG: dict[str, str] | None = None
_PATTERN_TO_TAG_FOLDED: dict[str, str] | None = None

# Llama 3.1 model name for fallback
LLAMA_MODEL = "llama3.1"

# ---------------------------------------------------------------------------
# Excel formatting helpers (shared with tc_to_excel.py)
# ---------------------------------------------------------------------------

_STEP_NO_RE = re.compile(r"\bA(\d+)\b", re.IGNORECASE)

_SECTION_FONT = Font(bold=True, color="001F4E78")
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="00EEF3FB")
_GROUP_FILLS = [
    PatternFill(fill_type="solid", fgColor="00FDE2E2"),
    PatternFill(fill_type="solid", fgColor="00E3F2FD"),
    PatternFill(fill_type="solid", fgColor="00E8F5E9"),
    PatternFill(fill_type="solid", fgColor="00E3F2FD"),
]
_RED_INLINE_FONT = InlineFont(color="00FF0000")


def _extract_group_step(path_text: str) -> str | None:
    """
    Extract the grouping step from a path string (mirrors tc_to_excel.py logic).
    - If path has "End", take the A<N> token immediately before "End".
    - Otherwise take the last A<N> token.
    """
    if not path_text:
        return None
    tokens = re.split(r"\s*(?:->|→|,)\s*", path_text.strip())
    tokens = [t.strip() for t in tokens if t.strip()]
    end_idx = next((i for i, t in enumerate(tokens) if t.lower() == "end"), -1)
    if end_idx != -1:
        if end_idx > 0:
            candidate = tokens[end_idx - 1]
            if _STEP_NO_RE.fullmatch(candidate):
                return candidate
        return None
    else:
        last = tokens[-1] if tokens else ""
        return last if _STEP_NO_RE.fullmatch(last) else None


def _step_sort_key(step_no: str) -> tuple[int, str]:
    m = _STEP_NO_RE.fullmatch(step_no.strip())
    return (int(m.group(1)), step_no) if m else (10**9, step_no)


def _auto_fit_column(ws: Worksheet, col_idx: int, extra: int = 2) -> None:
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            for part in str(cell.value).split("\n"):
                max_len = max(max_len, len(part))
    ws.column_dimensions[col_letter].width = max_len + extra


def _write_formatted_excel(
    ws,
    rows_data: list[dict[str, Any]],
    step_name_map: dict[str, str] | None = None,
) -> None:
    """
    Write all test-case rows into *ws* using the same layout/style as tc_to_excel.py.
    Each dict in rows_data must have keys:
        tc_id (original, ignored – we re-number),
        conditions, steps, bot_responses, path, expected_action,
        test_data, highlight_last_step (optional bool)
    """
    step_name_map = step_name_map or {}

    # Header row
    headers = ["TC_ID", "Conditions", "Test Scenario", "Bot Responses",
               "Path", "Expected Action Code", "Test Data"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Group cases by path step
    grouped: dict[str, list[dict[str, Any]]] = {}
    for case in rows_data:
        group_step = _extract_group_step(str(case.get("path", ""))) or "UNKNOWN"
        grouped.setdefault(group_step, []).append(case)

    ordered_groups = sorted(grouped.keys(), key=_step_sort_key)

    current_group_key: tuple[Any, ...] | None = None
    group_index = -1
    display_tc_index = 1
    row = 2

    for step_no in ordered_groups:
        step_name = step_name_map.get(step_no, step_no)
        section_label = f"{step_no} - {step_name}"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell_sec = ws.cell(row=row, column=1, value=section_label)
        cell_sec.font = _SECTION_FONT
        cell_sec.fill = _SECTION_FILL
        cell_sec.alignment = Alignment(wrap_text=True)
        row += 1

        for case in grouped[step_no]:
            tc_id = f"TC{display_tc_index:03d}"
            conditions = str(case.get("conditions", ""))
            steps: list[str] = case.get("steps", [])
            bot_responses: list[str] = case.get("bot_responses", [])
            expected_action = str(case.get("expected_action", case.get("expected_action_code", "N/A")))
            tc_path = str(case.get("path", ""))
            highlight_last_step = bool(case.get("highlight_last_step", False))
            test_data = str(case.get("test_data", ""))

            numbered_steps = [f"{i}. {s}" for i, s in enumerate(steps, 1)]
            scenario = "\n".join(numbered_steps)

            numbered_responses: list[str] = []
            for i, r in enumerate(bot_responses, 1):
                numbered_responses.append(f"{i}. {r}" if r else f"{i}. ")
            bot_responses_text = "\n".join(numbered_responses)

            # Determine fill colour based on logical grouping
            group_key = (conditions, tuple(steps), tc_path, expected_action)
            if group_key != current_group_key:
                group_index += 1
                current_group_key = group_key
            fill = _GROUP_FILLS[group_index % len(_GROUP_FILLS)]

            # Column A – TC_ID
            ws.cell(row=row, column=1, value=tc_id)

            # Column B – Conditions
            ws.cell(row=row, column=2, value=conditions)

            # Column C – Test Scenario (rich text if highlight_last_step)
            cell_scn = ws.cell(row=row, column=3)
            if highlight_last_step and numbered_steps:
                rt = CellRichText()
                if len(numbered_steps) > 1:
                    rt.append("\n".join(numbered_steps[:-1]) + "\n")
                rt.append(TextBlock(_RED_INLINE_FONT, numbered_steps[-1]))
                cell_scn.value = rt
            else:
                cell_scn.value = scenario
            cell_scn.alignment = Alignment(wrap_text=True)

            # Column D – Bot Responses (rich text if highlight_last_step)
            cell_bot = ws.cell(row=row, column=4)
            if highlight_last_step and numbered_responses:
                rt_bot = CellRichText()
                if len(numbered_responses) > 1:
                    rt_bot.append("\n".join(numbered_responses[:-1]) + "\n")
                rt_bot.append(TextBlock(_RED_INLINE_FONT, numbered_responses[-1]))
                cell_bot.value = rt_bot
            else:
                cell_bot.value = bot_responses_text
            cell_bot.alignment = Alignment(wrap_text=True)

            # Column E – Path
            ws.cell(row=row, column=5, value=tc_path)

            # Column F – Expected Action Code
            ws.cell(row=row, column=6, value=expected_action)

            # Column G – Test Data
            cell_g = ws.cell(row=row, column=7, value=test_data)
            cell_g.alignment = Alignment(wrap_text=True)

            # Apply fill to columns A-F (G stays default)
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill

            row += 1
            display_tc_index += 1

    # Auto-fit all columns
    for c in range(1, 8):
        _auto_fit_column(ws, c, extra=4)


# ---------------------------------------------------------------------------
# TLS Client Bot helpers
# ---------------------------------------------------------------------------

@contextmanager
def _temporary_chdir(path: Path):
    old = Path.cwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(old)


def _load_tls_intents() -> dict[str, Any]:
    global _TLS_INTENTS_DATA, _TAGS_SET, _PATTERN_TO_TAG, _PATTERN_TO_TAG_FOLDED
    if _TLS_INTENTS_DATA is not None:
        return _TLS_INTENTS_DATA

    if not TLS_INTENTS_PATH.exists():
        raise FileNotFoundError(f"Missing tls bot intents.json: {TLS_INTENTS_PATH}")

    data = json.loads(TLS_INTENTS_PATH.read_text(encoding="utf-8"))
    _TLS_INTENTS_DATA = data
    tags_set: set[str] = set()
    pattern_to_tag: dict[str, str] = {}
    pattern_to_tag_folded: dict[str, str] = {}

    for intent in data.get("intents", []):
        tag = str(intent.get("tag", "")).strip()
        if not tag:
            continue
        tags_set.add(tag)
        for pattern in intent.get("patterns", []) or []:
            p = str(pattern).strip().lower()
            if p and p not in pattern_to_tag:
                pattern_to_tag[p] = tag
            folded_p = _fold_for_match(p)
            if folded_p and folded_p not in pattern_to_tag_folded:
                pattern_to_tag_folded[folded_p] = tag

    _TAGS_SET = tags_set
    _PATTERN_TO_TAG = pattern_to_tag
    _PATTERN_TO_TAG_FOLDED = pattern_to_tag_folded
    return data


def _load_tls_get_response():
    global _TLS_GET_RESPONSE
    if _TLS_GET_RESPONSE is not None:
        return _TLS_GET_RESPONSE

    if not TLS_BOT_DIR.exists():
        _TLS_GET_RESPONSE = None
        return None

    try:
        tls_str = str(TLS_BOT_DIR)
        if tls_str not in sys.path:
            sys.path.insert(0, tls_str)

        with _temporary_chdir(TLS_BOT_DIR):
            from chat import get_response as tls_get_response  # type: ignore

        _TLS_GET_RESPONSE = tls_get_response
        return _TLS_GET_RESPONSE
    except Exception:
        _TLS_GET_RESPONSE = None
        return None


def _extract_latest_attempt(raw_step: str) -> tuple[str | None, str | None]:
    attempt_matches = re.finditer(r"\b(lần\s*(\d+))\b", raw_step, flags=re.IGNORECASE)
    attempts = []
    for match in attempt_matches:
        attempts.append({
            'full': match.group(1),
            'number': int(match.group(2)),
            'position': match.start()
        })
    if attempts:
        attempts.sort(key=lambda x: x['position'])
        latest = attempts[-1]
        return latest['full'], str(latest['number'])
    return None, None


def _normalize_step_to_bot_message(step: str) -> str:
    _load_tls_intents()
    tags_set = _TAGS_SET or set()

    raw = (step or "").strip()
    if not raw:
        return raw

    parts = [p.strip() for p in re.split(r"\s*\\\s*", raw) if p.strip()]
    first_part = parts[0] if parts else raw
    is_multi_intent = len(parts) > 1
    first_folded = _fold_for_match(first_part)

    if "im lang" in first_folded:
        for tag in tags_set:
            if "im lang" in _fold_for_match(tag):
                return tag

    if "khong nghe ro" in first_folded:
        for tag in tags_set:
            if "khong nghe ro" in _fold_for_match(tag):
                return tag

    if is_multi_intent:
        selected_intent = parts[0]
        attempt_full, _ = _extract_latest_attempt(raw)
        if attempt_full:
            selected_intent = re.sub(r"\s*lần\s*\d+\s*", "", selected_intent, flags=re.IGNORECASE).strip()
            return f"{selected_intent} {attempt_full}".strip()
        return first_part

    for tag in tags_set:
        tag_folded = _fold_for_match(tag)
        if tag_folded == first_folded:
            return tag
        tag_first = tag_folded.split("/")[0].strip()
        if tag_first == first_folded:
            return tag

    return first_part


_BANNED_CHARS_RE = re.compile(r"[\*\^\$\#\@\%\&\?\!]")
_TRAILING_PUNCT_RE = re.compile(r"[\.!\?]+$")
_MULTISPACE_RE = re.compile(r"\s+")


def _sanitize_utterance(text: str) -> str:
    if not text:
        return "(không sinh được)"

    t = str(text).strip()
    if t.upper() == "_SILENCE_":
        return "_SILENCE_"

    if "do not understand" in t.lower():
        return "(không sinh được)"

    t = t.replace(",", " ")
    t = _BANNED_CHARS_RE.sub("", t)
    t = t.replace("\n", " ")
    t = _TRAILING_PUNCT_RE.sub("", t).strip()
    t = t.lower()
    t = _MULTISPACE_RE.sub(" ", t).strip()

    words = t.split(" ")
    if len(words) > MAX_WORDS_PER_RESPONSE:
        t = " ".join(words[:MAX_WORDS_PER_RESPONSE]).strip()

    return t if t else "(không sinh được)"


def _fold_for_match(s: str) -> str:
    base = unicodedata.normalize("NFD", s.lower())
    return "".join(ch for ch in base if unicodedata.category(ch) != "Mn")


def _build_llama_prompt_for_step(
    step: str,
    step_number: int,
    total_steps: int,
    context: list[str],
    expected_bot_response: str = "",
) -> str:
    context_text = ""
    if context:
        context_lines = [f"{i}. {ctx}" for i, ctx in enumerate(context, 1)]
        context_text = "Các câu trước đó:\n" + "\n".join(context_lines) + "\n\n"

    bot_response_text = ""
    if expected_bot_response:
        bot_response_text = f"\nBot sẽ đáp: \"{expected_bot_response}\""

    return f"""Bạn là KHÁCH HÀNG (KH) đang nói chuyện TRỰC TIẾP qua điện thoại với tổng đài viên về KHOẢN VAY.

NHIỆM VỤ: chỉ tạo 1 câu lời thoại KH cho bước số {step_number} dựa trên kịch bản.

{context_text}Kịch bản bước {step_number}: {step}{bot_response_text}

YÊU CẦU BẮT BUỘC:
- Chỉ tạo DUY NHẤT 1 câu thoại KH đúng với mõi intent tương ứng
- CHỈ là lời KH nói, không phải mô tả tình huống
- KHÔNG được nhắc lại hoặc diễn giải kịch bản
- KHÔNG được viết kiểu: "khách hàng im lặng", "người thân nói", "cuộc gọi..."
- Nếu tình huống là "im lặng" → viết câu tự nhiên thể hiện không nghe rõ (vd: alo em không nghe rõ)
- Nếu "người thân nghe hộ" → vẫn phải nói như người đang nghe điện thoại (vd: chị là mẹ chau đang nghe)

FORMAT CỨNG:
- Chỉ output 1 câu duy nhất, không đánh số
- KHÔNG thêm tiêu đề, KHÔNG markdown, KHÔNG giải thích
- KHÔNG ký tự đặc biệt: *^$#@%&?!
- KHÔNG dấu chấm, chấm hỏi, chấm than cuối câu
- KHÔNG dấu phẩy
- viết thường toàn bộ
- tối đa {MAX_WORDS_PER_RESPONSE} từ

VÍ DỤ:
Kịch bản: KH im lặng lần 1
Bot sẽ đáp: "alo em nghe không rõ ạ"
Output: alo em nghe không rõ anh nói lại giúp em với

Đầu ra:"""


def _generate_single_with_llama(
    step: str,
    step_number: int,
    total_steps: int,
    context: list[str],
    expected_bot_response: str = "",
) -> str:
    if not step or ollama_chat is None:
        return "(không sinh được)"

    prompt = _build_llama_prompt_for_step(step, step_number, total_steps, context, expected_bot_response)
    try:
        response = ollama_chat(
            model=LLAMA_MODEL,
            messages=[{"role": "user", "content": prompt}],
        )
        content = (response.get("message") or {}).get("content") or ""
        content = content.strip()
        content = re.sub(r"^\s*\d+[.)]\s*", "", content)
        content = _sanitize_utterance(content)
        return content if content else "(không sinh được)"
    except Exception as e:
        print(f"    [WARN] Llama generation failed for step {step_number}: {e}")
        return "(không sinh được)"


def _fallback_response_for_message(msg: str) -> str:
    _load_tls_intents()
    intents = _TLS_INTENTS_DATA.get("intents", []) if _TLS_INTENTS_DATA else []
    tags_set = _TAGS_SET or set()
    pattern_to_tag = _PATTERN_TO_TAG or {}
    pattern_to_tag_folded = _PATTERN_TO_TAG_FOLDED or {}

    m = (msg or "").strip()
    if not m:
        return "(không sinh được)"

    m_lower = m.lower()
    m_folded = _fold_for_match(m_lower)
    m_without_attempt = re.sub(r'\s*lần\s*\d+\s*', ' ', m_lower).strip()
    m_without_attempt_folded = _fold_for_match(m_without_attempt)

    tag: str | None = None
    for t in tags_set:
        t_folded = _fold_for_match(t)
        t_without_attempt = re.sub(r'\s*lần\s*\d+\s*', '', t_folded).strip()
        if t_folded == m_folded or t_without_attempt == m_without_attempt_folded:
            tag = t
            break
        t_first = t_folded.split("/")[0].strip()
        if t_first == m_folded or t_first == m_without_attempt_folded:
            tag = t
            break

    if not tag:
        tag = pattern_to_tag.get(m_lower) or pattern_to_tag.get(m_without_attempt)
    if not tag:
        tag = pattern_to_tag_folded.get(m_folded) or pattern_to_tag_folded.get(m_without_attempt_folded)
    if not tag:
        return "(không sinh được)"

    for intent in intents:
        if str(intent.get("tag", "")).strip() == tag:
            responses = intent.get("responses") or []
            if responses:
                selected = random.choice([str(r) for r in responses if str(r).strip()])
                return selected
    return "(không sinh được)"


def _generate_step_with_tls_bot(step: str) -> str:
    tls_get_response = _load_tls_get_response()
    msg = _normalize_step_to_bot_message(step)

    if tls_get_response is None:
        resp = _fallback_response_for_message(msg)
        return _sanitize_utterance(resp)

    try:
        resp = tls_get_response(msg, context=None, token=None)
        if not resp or "do not understand" in str(resp).lower():
            resp = _fallback_response_for_message(msg)
    except Exception:
        resp = _fallback_response_for_message(msg)

    return _sanitize_utterance(resp)


def _pick_random_bot_response(bot_response_str: str) -> str:
    if not bot_response_str:
        return ""
    parts = [p.strip() for p in re.split(r"\s*\\\s*", bot_response_str) if p.strip()]
    return random.choice(parts) if parts else bot_response_str


def _is_a0_path(path_value: Any) -> bool:
    return str(path_value or "").strip() == "A0"


def _generate_hybrid(steps: list[str], bot_responses: list[str], tc_id: str) -> str:
    """Hybrid generation: TLS bot first, Llama fallback per failed step."""
    print(f"\n[DEBUG] Processing TC: {tc_id} with hybrid approach")

    utterances: list[str] = []
    tls_failed_steps: list[int] = []

    for i, step in enumerate(steps, 1):
        tls_result = _generate_step_with_tls_bot(step)
        utterances.append(tls_result)
        if tls_result == "(không sinh được)":
            tls_failed_steps.append(i - 1)
            print(f"  Step {i}: TLS bot failed -> will use Llama fallback")
        else:
            print(f"  Step {i}: TLS bot succeeded -> '{tls_result}'")

    for idx in tls_failed_steps:
        step_number = idx + 1
        step = steps[idx]
        raw_bot_resp = bot_responses[idx] if idx < len(bot_responses) else ""
        expected_bot_response = _pick_random_bot_response(raw_bot_resp)

        context = [
            utterances[j]
            for j in range(idx)
            if j < len(utterances) and utterances[j] != "(không sinh được)"
        ]

        print(f"  Step {step_number}: Generating with Llama (context: {len(context)} previous utterances)")
        llama_result = _generate_single_with_llama(step, step_number, len(steps), context, expected_bot_response)
        utterances[idx] = llama_result
        print(f"    Llama result: '{llama_result}'")

    result = "\n".join(f"{i+1}. {p}" for i, p in enumerate(utterances[:len(steps)]))

    tls_success = len(steps) - len(tls_failed_steps)
    print(f"  Summary: {tls_success}/{len(steps)} steps from TLS, {len(tls_failed_steps)} steps from Llama")
    return result


# ---------------------------------------------------------------------------
# Public entry-points
# ---------------------------------------------------------------------------

def fill_excel_from_json(
    json_path: Path,
    excel_path: Path,
    model: str = DEFAULT_MODEL,
    limit: int | None = None,
) -> None:
    """Load TC JSON, generate test data, write formatted Excel (one row per TC)."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of test cases.")

    rows_data: list[dict[str, Any]] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            continue
        tc_id = str(item.get("tc_id", f"TC_{i+1:03d}"))
        tc_path = str(item.get("path", ""))

        steps = item.get("steps", [])
        if not isinstance(steps, Iterable):
            steps = []
        bot_responses = item.get("bot_responses", [])
        if not isinstance(bot_responses, Iterable):
            bot_responses = []

        clean_steps = [
            str(s).strip()
            for s in steps
            if str(s).strip() and not str(s).strip().startswith("(cycle to")
        ]
        clean_bot_responses = [str(r).strip() for r in bot_responses]
        while len(clean_bot_responses) < len(clean_steps):
            clean_bot_responses.append("")

        # A0 paths (or any case with zero steps) get blank test data
        if not clean_steps:
            test_data = ""
        else:
            test_data = _generate_hybrid(clean_steps, clean_bot_responses, tc_id)
            print(f"  Generated: {tc_id} ({len(clean_steps)} steps)")

        rows_data.append({
            "tc_id": tc_id,
            "conditions": str(item.get("conditions", "")),
            "steps": clean_steps,
            "bot_responses": clean_bot_responses,
            "expected_action": str(item.get("expected_action_code", "N/A")),
            "path": tc_path,
            "highlight_last_step": bool(item.get("highlight_last_step", False)),
            "test_data": test_data,
        })

    if limit is not None:
        rows_data = rows_data[:limit]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestCases"
    _write_formatted_excel(ws, rows_data)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def fill_excel_from_excel(
    excel_path: Path,
    out_path: Path,
    model: str = DEFAULT_MODEL,
    limit: int | None = None,
) -> None:
    """Read existing Excel, fill Test Data column, write new formatted Excel."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    max_row = ws.max_row
    if limit is not None:
        max_row = min(max_row, 1 + limit)

    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val is None:
            continue
        header_to_col[str(header_val).strip()] = col

    tc_id_col = header_to_col.get("TC_ID", 1)
    conditions_col = header_to_col.get("Conditions", 0)
    scenario_col = header_to_col.get("Test Scenario", 2)
    bot_responses_col = header_to_col.get("Bot Responses", 3)
    path_col = header_to_col.get("Path", 4)
    expected_action_col = header_to_col.get("Expected Action Code", 5)

    numbered_re = re.compile(r"^\s*(\d+)[.)]\s*(.*)$")

    def _parse_numbered_lines(cell_value: str) -> list[str]:
        result: list[str] = []
        for raw_line in str(cell_value).strip().split("\n"):
            line = raw_line.strip()
            if not line:
                continue
            m = numbered_re.match(raw_line)
            if m:
                cleaned = (m.group(2) or "").strip()
                result.append(cleaned)
            elif result:
                result[-1] = f"{result[-1].rstrip()} {line}".strip()
            else:
                result.append(line)
        return result

    rows_data: list[dict[str, Any]] = []
    for row in range(2, max_row + 1):
        tc_id = ws.cell(row=row, column=tc_id_col).value or ""
        path_val = str(ws.cell(row=row, column=path_col).value or "")

        scenario_cell = ws.cell(row=row, column=scenario_col).value or ""
        bot_responses_cell = (
            ws.cell(row=row, column=bot_responses_col).value or ""
            if bot_responses_col else ""
        )
        expected_action = ws.cell(row=row, column=expected_action_col).value or ""
        conditions = (
            ws.cell(row=row, column=conditions_col).value or ""
            if conditions_col else ""
        )

        steps = _parse_numbered_lines(str(scenario_cell))
        bot_responses = _parse_numbered_lines(str(bot_responses_cell))
        while len(bot_responses) < len(steps):
            bot_responses.append("")

        if not steps:
            test_data = ""
        else:
            test_data = _generate_hybrid(steps, bot_responses, str(tc_id))
            print(f"  Generated: {tc_id} ({len(steps)} steps)")

        rows_data.append({
            "tc_id": tc_id,
            "conditions": str(conditions),
            "steps": steps,
            "bot_responses": bot_responses,
            "expected_action": str(expected_action),
            "path": path_val,
            "highlight_last_step": False,
            "test_data": test_data,
        })

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "TestCases"
    _write_formatted_excel(new_ws, rows_data)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    new_wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate test data using hybrid approach: TLS Client Bot + Llama 3.1 fallback per step."
    )
    parser.add_argument(
        "--in",
        dest="in_path",
        type=Path,
        required=True,
        help="Input: JSON (tcgen output) or Excel (tc_to_excel output).",
    )
    parser.add_argument(
        "--out",
        dest="out_path",
        type=Path,
        default=None,
        help="Output Excel path. If input is Excel, defaults to overwriting input.",
    )
    parser.add_argument(
        "--model",
        type=str,
        default=DEFAULT_MODEL,
        help=f"Kept for compatibility (default: {DEFAULT_MODEL}).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N test cases (for testing).",
    )
    args = parser.parse_args()

    in_path = args.in_path
    if not in_path.exists():
        print(f"Error: Input file not found: {in_path}")
        return 1

    out_path = args.out_path
    if out_path is None:
        out_path = (
            in_path if in_path.suffix.lower() == ".xlsx"
            else Path("output/test_cases.xlsx")
        )

    print("=" * 60)
    print("HYBRID GENERATION MODE: TLS Client Bot + Llama 3.1 Fallback (per-step)")
    print("Bot responses may have multiple variants (separated by ' \\ ') -> random selection.")
    print("Including A0 test cases with empty test data.")
    print("=" * 60)

    if ollama_chat is None:
        print("⚠️  Warning: Ollama not installed. Llama fallback will not work.")
        print("   Install with: pip install ollama")
        print("   Or run: ollama pull llama3.1")
        print()

    if in_path.suffix.lower() == ".json":
        print(f"Reading JSON: {in_path}")
        print(f"Writing Excel with test data: {out_path}")
        fill_excel_from_json(in_path, out_path, model=args.model, limit=args.limit)
    else:
        print(f"Reading Excel: {in_path}")
        fill_excel_from_excel(in_path, out_path, model=args.model, limit=args.limit)

    print(f"\nDone. Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
