from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Iterable, List

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants for header styling
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="008EA9DB", end_color="008EA9DB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="00000000")
ROUND_HEADER_FILL = PatternFill(start_color="00E2EFDA", end_color="00E2EFDA", fill_type="solid")
ROUND_HEADER_FONT = Font(bold=True, color="00000000")

# Main headers (fixed part, columns A–F only – column G becomes part of Data collection)
FIXED_MAIN_HEADERS = [
    "TC_ID",
    "Flow",
    "Conditions",
    "Data input",
    "Test Scenario",
    "Expected Bot Responses",
]

# Sub‑headers for each round (8 columns each) — used by all pipelines except multi_responses
ROUND_SUB_HEADERS = [
    "Tester",
    "User Message",
    "Call Id",
    "Test Results\n(Bot responses)",
    "Test Results\n(data collection)",
    "Error Type",
    "Error Description",
    "FPT Comment",
]

# Sub‑headers for the multi_responses pipeline (7 columns, single Test Results)
ROUND_SUB_HEADERS_MULTI = [
    "Tester",
    "User Message",
    "Call Id",
    "Test Results",
    "Error Type",
    "Error Description",
    "FPT Comment",
]

# Section header style (for grouping rows)
SECTION_FONT = Font(bold=True, color="000000")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="ffff00")

# Group fill colors – only used when use_group_fills=True, applied to columns A-G
GROUP_FILLS = [
    PatternFill(fill_type="solid", fgColor="00FDE2E2"),
    PatternFill(fill_type="solid", fgColor="00E3F2FD"),
    PatternFill(fill_type="solid", fgColor="00E8F5E9"),
    PatternFill(fill_type="solid", fgColor="00E3F2FD"),
]

RED_INLINE_FONT = InlineFont(color="00FF0000")

# Border style
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Base column widths (indices relative to fixed columns)
BASE_COL_WIDTHS = {
    1: 10,   # TC_ID
    2: 25,   # Flow (Path)
    3: 40,   # Conditions
    4: 30,   # Data input
    5: 40,   # Test Scenario
    6: 90,   # Expected Bot Responses
}
# Default width for dynamic columns
DYN_COL_WIDTH = 20
# Round sub-column widths (same for both rounds)
ROUND_COL_WIDTHS = [15, 40, 20, 20, 20, 20, 20, 20]  # Tester, Test Data, Call Id, Test Results (Bot), Test Results (data), Error Type, Error Description, FPT Comment
ROUND_COL_WIDTHS_MULTI = [15, 40, 20, 20, 20, 20, 20]  # Tester, Test Data, Call Id, Test Results, Error Type, Error Description, FPT Comment

INPUT_ROW_COLUMNS: list[tuple[str, str]] = [
    ("Step no", "step_no"),
    ("Step name", "step_name"),
    ("Conditions", "conditions"),
    ("Customer intent", "customer_intent"),
    ("Bot response", "bot_response"),
    ("Bot response 2", "bot_response_2"),
    ("Bot response 3", "bot_response_3"),
    ("Bot response 4", "bot_response_4"),
    ("Bot response 5", "bot_response_5"),
    ("Next Step", "next_step"),
    ("Action code", "action_code"),
]

INPUT_ROW_KEYS = {key for _, key in INPUT_ROW_COLUMNS}

# Conditional formatting colors for Test Results columns.
RESULT_COLORS = {
    "Pass": "00C6EFCE",        # light green
    "Fail": "00FFC7CE",        # light red
    "Pending": "00FFEB9C",     # light yellow
    "Fixed": "00FFEB9C",
    "Un-executed": "00FFEB9C",
    "Need review": "00D8BFD8", # light purple
    "Todo": "00FFFFFF",        # white
}
ERROR_TYPE_COLORS = {
    "ASR Mistake": "00FFFFFF",
    "Intent": "00FFFFFF",
    "Flow": "00FFFFFF",
    "Script": "00FFFFFF",
    "Other": "00FFFFFF",
}


def _load_testcases(path: Path) -> list[dict[str, Any]]:
    """Read a JSON array of test cases and return a list of cleaned dicts."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of test cases.")

    cases: list[dict[str, Any]] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Item at index {i} is not an object.")

        tc_id = str(item.get("tc_id", f"TC{i+1}"))
        conditions = str(item.get("conditions", ""))
        steps = item.get("steps", [])
        bot_responses = item.get("bot_responses", [])
        expected_action_code = item.get("expected_action_code", "N/A")
        tc_path = str(item.get("path", ""))
        highlight_last_step = bool(item.get("highlight_last_step", False))
        test_data = str(item.get("test_data", ""))
        source_columns = item.get("source_columns", {})

        if not isinstance(steps, Iterable):
            raise ValueError(f"'steps' for {tc_id} must be a list.")
        if not isinstance(bot_responses, Iterable):
            raise ValueError(f"'bot_responses' for {tc_id} must be a list.")

        # Remove cycle markers like "(cycle to A1)"
        clean_steps: list[str] = []
        for s in steps:
            s_str = str(s)
            if s_str.strip().startswith("(cycle to"):
                continue
            clean_steps.append(s_str)

        clean_bot_responses: list[str] = [str(r) for r in bot_responses]
        while len(clean_bot_responses) < len(clean_steps):
            clean_bot_responses.append("")

        cases.append({
            "tc_id": tc_id,
            "conditions": conditions,
            "steps": clean_steps,
            "bot_responses": clean_bot_responses,
            "expected_action_code": expected_action_code,
            "path": tc_path,
            "highlight_last_step": highlight_last_step,
            "test_data": test_data,
            "source_columns": source_columns,
        })
    return cases


def _auto_fit_column(ws: Worksheet, col_idx: int, extra: int = 2) -> None:
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            for part in str(cell.value).split("\n"):
                max_len = max(max_len, len(part))
    ws.column_dimensions[col_letter].width = max_len + extra


def _auto_fit_row_heights(ws: Worksheet, min_height: float = 15.0, line_height: float = 15.0) -> None:
    """Adjust row heights based on line count of the Expected Bot Responses column (col 6)."""
    for row_cells in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=6, max_col=6):
        cell = row_cells[0]
        lines = len(str(cell.value).split("\n")) if cell.value else 1
        ws.row_dimensions[cell.row].height = max(lines * line_height, min_height)


_STEP_NO_RE = re.compile(r"A(\d+)(?:\.(\d+))?", re.IGNORECASE)


def _extract_group_step(path_text: str) -> str | None:
    """Extract the grouping step from a path string.
    - If path contains "End", return the A<N> token immediately before "End".
    - Otherwise return the last A<N> token.
    """
    if not path_text:
        return None
    tokens = re.split(r"\s*(?:->|→|,)\s*", path_text.strip())
    tokens = [t.strip() for t in tokens if t.strip()]
    end_idx = next((i for i, t in enumerate(tokens) if t.lower() == "end"), -1)
    if end_idx != -1:
        if end_idx > 0:
            candidate = tokens[end_idx - 1]
            if _STEP_NO_RE.fullmatch(candidate):
                return candidate
        return None
    else:
        last = tokens[-1] if tokens else ""
        return last if _STEP_NO_RE.fullmatch(last) else None


def _step_sort_key(step_no: str) -> tuple[int, int, str]:
    m = _STEP_NO_RE.fullmatch(step_no.strip())
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) is not None else 0
        return (major, minor, step_no)
    return (10**9, 10**9, step_no)


def _load_step_name_map(testcases_path: Path, rows_path: Path | None = None) -> dict[str, str]:
    """Try to load step_no → step_name mapping from a rows JSON file."""
    candidates: list[Path] = []
    if rows_path is not None and rows_path.exists():
        candidates.append(rows_path)
    in_name = testcases_path.name
    if in_name.startswith("testcases_"):
        candidates.append(testcases_path.with_name(in_name.replace("testcases_", "rows_", 1)))
    candidates.append(testcases_path.with_name("rows_multi_vb.json"))
    candidates.append(testcases_path.with_name("rows_e2e_vb.json"))

    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            data = json.loads(candidate.read_text(encoding="utf-8"))
            if not isinstance(data, list):
                continue
            mapping: dict[str, str] = {}
            for item in data:
                if not isinstance(item, dict):
                    continue
                step_no = str(item.get("step_no", "")).strip()
                step_name = str(item.get("step_name", "")).strip()
                if step_no and step_name and step_name.upper() != step_no.upper() and step_no not in mapping:
                    mapping[step_no] = step_name
            if mapping:
                return mapping
        except Exception:
            continue
    return {}


def _load_source_rows(rows_path: Path | None) -> list[dict[str, Any]] | None:
    if rows_path is None or not rows_path.exists():
        return None

    data = json.loads(rows_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"Rows JSON must be a list: {rows_path}")

    rows: list[dict[str, Any]] = []
    for idx, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Rows JSON item at index {idx} is not an object: {rows_path}")
        rows.append(item)
    return rows


def _apply_borders(ws: Worksheet) -> None:
    """Apply thin borders to all used cells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def _set_column_widths(ws: Worksheet, dynamic_keys: List[str], is_multi_responses: bool = False) -> None:
    """Set column widths based on BASE_COL_WIDTHS, dynamic keys, and round columns."""
    # Fixed columns 1-6
    for col_idx, width in BASE_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if is_multi_responses:
        # Round columns start immediately at column 7 (no data collection section)
        for i, width in enumerate(ROUND_COL_WIDTHS_MULTI):
            ws.column_dimensions[get_column_letter(7 + i)].width = width
    else:
        # Dynamic columns start immediately after fixed columns A-F.
        for i, key in enumerate(dynamic_keys, start=7):
            ws.column_dimensions[get_column_letter(i)].width = DYN_COL_WIDTH
        # Round 1 columns
        round1_start = 7 + len(dynamic_keys)
        for i, width in enumerate(ROUND_COL_WIDTHS):
            ws.column_dimensions[get_column_letter(round1_start + i)].width = width


def _apply_alignments(ws: Worksheet, dynamic_keys: List[str], section_header_rows: list[int] | None = None, is_multi_responses: bool = False) -> None:
    """Vertical center for all cells, horizontal center for column A, round results/error columns, and header rows.
    Section header rows (if provided) are set to left alignment afterwards."""
    if is_multi_responses:
        # round starts at column 7, Test Results at +3, Error Type at +4
        center_cols = {7 + 3, 7 + 4}  # Test Results, Error Type
    else:
        round1_start = 7 + len(dynamic_keys)
        center_cols = {round1_start + 3, round1_start + 4, round1_start + 5}  # Results (Bot), Results (data), Error Type

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.column == 1 or cell.row == 2 or cell.row == 3 or cell.column in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Override section header rows to left alignment
    if section_header_rows:
        for r in section_header_rows:
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _add_data_validations(ws: Worksheet, dynamic_keys: List[str], max_row: int, is_multi_responses: bool = False) -> None:
    """Add dropdown lists to Test Results and Error Type columns."""
    data_start_row = 4  # row 1=Ngày test, row 2=main headers, row 3=sub-headers

    if is_multi_responses:
        result_cols = [7 + 3]        # single Test Results column
        error_col = 7 + 4            # Error Type
    else:
        round1_start = 7 + len(dynamic_keys)
        result_cols = [round1_start + 3, round1_start + 4]  # Bot responses + data collection
        error_col = round1_start + 5

    # Test Results list
    dv_results = DataValidation(
        type="list",
        formula1='"Todo,Pass,Fail,Pending,Fixed,Un-executed,Need review"',
        allow_blank=True,
    )
    dv_results.error = "Please select a value from the list."
    dv_results.errorTitle = "Invalid Result"
    ws.add_data_validation(dv_results)
    for col in result_cols:
        dv_results.add(f"{get_column_letter(col)}{data_start_row}:{get_column_letter(col)}{max_row}")

    # Error Type list
    dv_error = DataValidation(
        type="list",
        formula1='"ASR Mistake,Intent,Flow,Script,Other"',
        allow_blank=True,
    )
    dv_error.error = "Please select a value from the list."
    dv_error.errorTitle = "Invalid Error Type"
    ws.add_data_validation(dv_error)
    dv_error.add(f"{get_column_letter(error_col)}{data_start_row}:{get_column_letter(error_col)}{max_row}")


def _add_conditional_formatting(ws: Worksheet, dynamic_keys: List[str], max_row: int, is_multi_responses: bool = False) -> None:
    """Apply background color rules to Test Results and Error Type columns."""
    data_start_row = 4  # row 1=Ngày test, row 2=main headers, row 3=sub-headers

    if is_multi_responses:
        result_cols = [7 + 3]   # single Test Results column
        error_col = 7 + 4       # Error Type
    else:
        round1_start = 7 + len(dynamic_keys)
        result_cols = [round1_start + 3, round1_start + 4]  # Bot responses + data collection
        error_col = round1_start + 5

    # Test Results columns
    for result_col in result_cols:
        for value, color in RESULT_COLORS.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            dxf = DifferentialStyle(fill=fill)
            rule = Rule(
                type="cellIs",
                operator="equal",
                formula=[f'"{value}"'],
                dxf=dxf,
            )
            ws.conditional_formatting.add(f"{get_column_letter(result_col)}{data_start_row}:{get_column_letter(result_col)}{max_row}", rule)

    # Error Type column
    for value, color in ERROR_TYPE_COLORS.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(
            type="cellIs",
            operator="equal",
            formula=[f'"{value}"'],
            dxf=dxf,
        )
        ws.conditional_formatting.add(f"{get_column_letter(error_col)}{data_start_row}:{get_column_letter(error_col)}{max_row}", rule)


def _title_from_key(key: str) -> str:
    return " ".join(part for part in key.split("_") if part).title()

_CHECKLIST_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    (
        "Tốc độ xử lý của BOT",
        [
            ("Thời gian BOT trả về câu trả lời nhanh (<= 3s)", "Bắt buộc"),
        ],
    ),
    (
        "Đáp ứng tính chính xác câu trả lời",
        [],
    ),
    (
        "Đi đúng luồng và đáp ứng nhu cầu hoạt động (Pass >90% Test Case)",
        [
            ("Đi đúng luồng như mô tả kịch bản >= 90%", "Bắt buộc"),
            ("Đi đúng luồng mô tả kịch bản >= 90%", "Bắt buộc"),
        ],
    ),
    (
        "Đáp ứng giao diện người dùng",
        [
            ("Chính tả, font chữ, backgroud, button…", "Bắt buộc"),
            ("Câu trả lời đúng trọng tâm, dễ hiểu, thuyết phục người dùng…", "Bắt buộc"),
        ],
    ),
]


def _populate_checklist_ws(ws: "Worksheet") -> None:  # noqa: F821
    """Fill an existing worksheet with the FPT BOT evaluation checklist content."""

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 20

    YELLOW_FILL = PatternFill(fill_type="solid", fgColor="00FFFF00")
    COUNTER_BOX_FILL = PatternFill(fill_type="solid", fgColor="00E0E0E0")
    ITALIC_GRAY = Font(italic=True, color="00808080", size=9)

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="Tiêu Chí Đánh Giá BOT")
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Rows 2-7: info labels (left) + Passed/Failed counters (right) ─────────
    for row_num, label in [
        (2, "Tên dự án:"),
        (3, "Người kiểm tra:"),
        (4, "Ngày kiểm tra:"),
        (6, "Người kiểm tra cuối cùng:"),
        (7, "Ngày kiểm tra cuối cùng:"),
    ]:
        ws.cell(row=row_num, column=1, value=label)
        ws.row_dimensions[row_num].height = 18
    ws.row_dimensions[5].height = 8

    # Summary labels (right side, rows 2-3)
    ws.merge_cells("E2:F2")
    lbl_p = ws.cell(row=2, column=5, value='Số mục "Passed" :')
    lbl_p.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("E3:F3")
    lbl_f = ws.cell(row=3, column=5, value='Số mục "Failed" :')
    lbl_f.alignment = Alignment(horizontal="right", vertical="center")

    # COUNTIF counter cells (row 2-3, col G) — reference D10:D500
    cnt_p = ws.cell(row=2, column=7, value='=COUNTIF(D10:D500,"Pass")')
    cnt_p.font = Font(bold=True, size=16)
    cnt_p.fill = COUNTER_BOX_FILL
    cnt_p.alignment = Alignment(horizontal="center", vertical="center")
    cnt_p.border = THIN_BORDER

    cnt_f = ws.cell(row=3, column=7, value='=COUNTIF(D10:D500,"Fail")')
    cnt_f.font = Font(bold=True, size=16, color="00FF0000")
    cnt_f.fill = COUNTER_BOX_FILL
    cnt_f.alignment = Alignment(horizontal="center", vertical="center")
    cnt_f.border = THIN_BORDER

    # ── Row 8: blank gap ──────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 6

    # ── Row 9: table header ───────────────────────────────────────────────────
    TABLE_HDR = 9
    ws.row_dimensions[TABLE_HDR].height = 40
    for col_idx, text in enumerate(
        [
            "STT",
            "Các mục kiểm tra",
            "Danh mục",
            'Đánh giá\n(passed/failed/"N/A")',
            "Ghi chú",
            "Độ ưu tiên",
            "Hướng dẫn",
        ],
        start=1,
    ):
        c = ws.cell(row=TABLE_HDR, column=col_idx, value=text)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

    # ── Data rows (start at row 10) ───────────────────────────────────────────
    row = TABLE_HDR + 1
    stt = 1

    for cat_name, items in _CHECKLIST_SECTIONS:
        # Yellow category header — merged A:G
        ws.merge_cells(f"A{row}:G{row}")
        cc = ws.cell(row=row, column=1, value=cat_name)
        cc.fill = YELLOW_FILL
        cc.font = Font(bold=True)
        cc.alignment = Alignment(horizontal="left", vertical="center")
        cc.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        for item_text, do_uu_tien in items:
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=1, value=stt).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=row, column=2, value=item_text).alignment = Alignment(
                vertical="center", wrap_text=True
            )
            dg = ws.cell(row=row, column=4, value="")
            dg.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=6, value=do_uu_tien).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.row_dimensions[row].height = 18
            stt += 1
            row += 1

        # Italic placeholder row
        ws.merge_cells(f"A{row}:G{row}")
        ph = ws.cell(
            row=row,
            column=1,
            value="<Thêm các câu hỏi khác phía trên hàng này nếu có>",
        )
        ph.font = ITALIC_GRAY
        ph.alignment = Alignment(horizontal="left", vertical="center")
        ph.border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Outer medium border around the entire table block (row 9 → last section row) ──
    _OUTER = Side(style="medium")
    _THIN = Side(style="thin")
    table_end_row = row - 1

    for c in range(1, 8):
        # Top edge (table header row)
        tc = ws.cell(row=TABLE_HDR, column=c)
        tc.border = Border(
            top=_OUTER,
            bottom=_THIN,
            left=_OUTER if c == 1 else _THIN,
            right=_OUTER if c == 7 else _THIN,
        )
        # Bottom edge (last table row)
        bc = ws.cell(row=table_end_row, column=c)
        bc.border = Border(
            top=_THIN,
            bottom=_OUTER,
            left=_OUTER if c == 1 else _THIN,
            right=_OUTER if c == 7 else _THIN,
        )
    for r in range(TABLE_HDR + 1, table_end_row):
        lc = ws.cell(row=r, column=1)
        lc.border = Border(top=_THIN, bottom=_THIN, left=_OUTER, right=_THIN)
        rc = ws.cell(row=r, column=7)
        rc.border = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_OUTER)

    # Dropdown validation for all Đánh giá cells (D10:D{row-1})
    dv = DataValidation(type="list", formula1='"Pass,Fail,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D10:D{row - 1}")

    # Conditional formatting: green for Pass, red for Fail
    for value, hex_color in [("Pass", "00C6EFCE"), ("Fail", "00FFC7CE")]:
        dxf = DifferentialStyle(fill=PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid"))
        ws.conditional_formatting.add(
            f"D10:D{row - 1}",
            Rule(type="cellIs", operator="equal", formula=[f'"{value}"'], dxf=dxf),
        )

    # ── Footer ────────────────────────────────────────────────────────────────
    row += 1
    nhx = ws.cell(row=row, column=1, value="*Nhận xét")
    nhx.font = Font(bold=True)
    ws.row_dimensions[row].height = 18

    comment_start = row + 1
    comment_end = row + 3
    ws.merge_cells(f"A{comment_start}:G{comment_end}")
    box = ws.cell(row=comment_start, column=1)
    box.border = THIN_BORDER
    for r in range(comment_start, comment_end + 1):
        ws.row_dimensions[r].height = 18
    row = comment_end + 2

    gy = ws.cell(row=row, column=1, value="* Gợi ý")
    gy.font = Font(bold=True)
    ws.row_dimensions[row].height = 18
    row += 1
    for note in ["[    ] - Pass", "[    ] - Fail", "[    ] - N/A"]:
        ws.cell(row=row, column=1, value=note)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    nc = ws.cell(
        row=row,
        column=1,
        value="Chú ý: Đánh giá có thể tùy theo từng dự án, tùy yêu cầu đơn vị, khách hàng",
    )
    nc.font = Font(italic=True, size=9)
    ws.merge_cells(f"A{row}:G{row}")


def _add_checklist_sheet(wb: openpyxl.Workbook) -> None:
    """Create a 'Checklist' sheet as the next sheet in the workbook."""
    ws = wb.create_sheet("Checklist")
    _populate_checklist_ws(ws)


def _add_data_input_sheet(
    wb: openpyxl.Workbook,
    schema_keys: list[str] | None = None,
    schema_value_rows: list[list[str]] | None = None,
) -> None:
    """Create a 'Data input' sheet with columns ID, Conditions, Data plus schema key columns."""
    ws = wb.create_sheet("Data input")
    extra_keys = schema_keys or []
    headers = ["Data ID", "Conditions", "Data"] + extra_keys
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    for i in range(len(extra_keys)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 25

    if schema_value_rows:
        for row_idx, row_data in enumerate(schema_value_rows, start=2):
            for col_offset, value in enumerate(row_data):
                if value:
                    ws.cell(row=row_idx, column=4 + col_offset, value=value)

    _apply_borders(ws)


def _add_data_schema_sheet(
    wb: openpyxl.Workbook,
    schema_all_rows: list[list[str]] | None = None,
) -> None:
    """Create a 'Data Schema' sheet from input schema data, or with default structure."""
    ws = wb.create_sheet("Data Schema")

    if schema_all_rows:
        for row_idx, row_data in enumerate(schema_all_rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                if value:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        max_col = ws.max_column or 5
        for col_idx in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    else:
        headers = ["ID", "Field", "Type", "Description", "Note"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30

    _apply_borders(ws)


def export_to_excel(
    cases: list[dict[str, Any]],
    out_path: Path,
    step_name_map: dict[str, str] | None = None,
    *,
    group_by_step: bool = False,
    use_group_fills: bool = False,
    allow_highlight_last: bool = False,
    is_multi_responses: bool = False,
    source_rows: list[dict[str, Any]] | None = None,
    wb: openpyxl.Workbook | None = None,
    sheet_name: str = "TestCases",
    schema_keys: list[str] | None = None,
    schema_value_rows: list[list[str]] | None = None,
    schema_all_rows: list[list[str]] | None = None,
) -> openpyxl.Workbook:
    """Write test cases to an Excel file with full formatting."""
    step_name_map = step_name_map or {}

    # Collect all dynamic keys from all test cases
    all_dynamic_keys = set()
    for case in cases:
        source_columns = case.get("source_columns", {})
        for k in source_columns:
            if k not in INPUT_ROW_KEYS:
                all_dynamic_keys.add(k)
    dynamic_keys = sorted(all_dynamic_keys)  # deterministic order

    _standalone = wb is None
    if _standalone:
        wb = openpyxl.Workbook()
        # Use the default sheet for Checklist (first sheet), then add test cases sheet
        wb.active.title = "Checklist"
        _populate_checklist_ws(wb.active)
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Calculate column ranges
    num_dynamic = len(dynamic_keys)
    if is_multi_responses:
        round1_start = 7                    # round columns start immediately after F
        active_sub_headers = ROUND_SUB_HEADERS_MULTI
        active_round_widths = ROUND_COL_WIDTHS_MULTI
    else:
        round1_start = 7 + num_dynamic      # column after dynamic data-collection columns
        active_sub_headers = ROUND_SUB_HEADERS
        active_round_widths = ROUND_COL_WIDTHS
    total_columns = round1_start + len(active_round_widths) - 1

    # ---------- Write row 1 ("Ngày test:" info row) ----------
    cell_date = ws.cell(row=1, column=6, value="Ngày test:")
    cell_date.alignment = Alignment(horizontal="left", vertical="center")

    # ---------- Write row 2 (top header row) ----------
    # 1. Fixed headers A-F
    for col_idx, header_text in enumerate(FIXED_MAIN_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header_text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 2. "Expected Data Collection" group — skipped for multi_responses
    if not is_multi_responses and num_dynamic > 0:
        data_collection_start = 7
        data_collection_end = 6 + num_dynamic
        ws.merge_cells(start_row=2, start_column=data_collection_start, end_row=2, end_column=data_collection_end)
        cell_data = ws.cell(row=2, column=data_collection_start, value="Expected Data Collection")
        cell_data.fill = HEADER_FILL
        cell_data.font = HEADER_FONT
        cell_data.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3. Round 1 merged header
    ws.merge_cells(start_row=2, start_column=round1_start, end_row=2, end_column=round1_start + len(active_round_widths) - 1)
    cell_r1 = ws.cell(row=2, column=round1_start, value="Round 01")
    cell_r1.fill = ROUND_HEADER_FILL
    cell_r1.font = ROUND_HEADER_FONT
    cell_r1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---------- Write row 3 (sub-header row) ----------
    if not is_multi_responses:
        # Dynamic columns start at column G. The internal action_code field is not displayed.
        for i, key in enumerate(dynamic_keys, start=7):
            cell = ws.cell(row=3, column=i, value=key)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Round 1 sub-headers
    for i, sub_header in enumerate(active_sub_headers):
        col_idx = round1_start + i
        cell = ws.cell(row=3, column=col_idx, value=sub_header)
        cell.fill = ROUND_HEADER_FILL
        cell.font = ROUND_HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Merge fixed columns A–F vertically (rows 2 and 3)
    for col_idx in range(1, 7):
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)

    # Dynamic columns stay as individual cells under the data-collection group.

    # ---------- Common helper for writing a data row ----------
    def _write_data_row(
        row: int, case: dict[str, Any], display_tc_index: int, fill: PatternFill | None = None
    ) -> int:
        """Write a single data row (columns A through total_columns) and return the next row index."""
        tc_id = f"TC{display_tc_index:03d}"
        conditions = str(case.get("conditions", ""))
        steps: list[str] = case.get("steps", [])
        bot_responses: list[str] = case.get("bot_responses", [])
        tc_path = str(case.get("path", ""))
        highlight_last = bool(case.get("highlight_last_step", False))
        test_data = str(case.get("test_data", ""))
        source_columns = case.get("source_columns", {})

        numbered_steps = [f"{i}. {s}" for i, s in enumerate(steps, 1)]
        scenario = "\n".join(numbered_steps)

        numbered_responses: list[str] = []
        for i, r in enumerate(bot_responses, 1):
            numbered_responses.append(f"{i}. {r}" if r else f"{i}. ")
        bot_responses_text = "\n".join(numbered_responses)

        # For multi_responses: auto-enable highlight for TCs whose flow path is A0
        if is_multi_responses and not highlight_last:
            highlight_last = _extract_group_step(tc_path) == "A0"

        # Column A – TC_ID
        ws.cell(row=row, column=1, value=tc_id)
        # Column B – Flow (Path)
        ws.cell(row=row, column=2, value=tc_path)
        # Column C – Conditions
        ws.cell(row=row, column=3, value=conditions)
        # Column D – Data input (kept empty)
        ws.cell(row=row, column=4, value="")
        # Column E – Test Scenario (red highlight only for non-multi pipelines)
        cell_scn = ws.cell(row=row, column=5)
        if allow_highlight_last and highlight_last and numbered_steps and not is_multi_responses:
            rt = CellRichText()
            if len(numbered_steps) > 1:
                rt.append("\n".join(numbered_steps[:-1]) + "\n")
            rt.append(TextBlock(RED_INLINE_FONT, numbered_steps[-1]))
            cell_scn.value = rt
        else:
            cell_scn.value = scenario
        # Column F – Expected Bot Responses (red highlight for last response in multi_responses)
        cell_bot = ws.cell(row=row, column=6)
        if allow_highlight_last and highlight_last and numbered_responses:
            rt_bot = CellRichText()
            if len(numbered_responses) > 1:
                rt_bot.append("\n".join(numbered_responses[:-1]) + "\n")
            rt_bot.append(TextBlock(RED_INLINE_FONT, numbered_responses[-1]))
            cell_bot.value = rt_bot
        else:
            cell_bot.value = bot_responses_text

        if not is_multi_responses:
            for i, key in enumerate(dynamic_keys):
                col_idx = 7 + i
                value = str(source_columns.get(key, ""))
                ws.cell(row=row, column=col_idx, value=value)

        if is_multi_responses:
            # Round 1 sub‑columns: Tester | Test Data | Call Id | Test Results | Error Type | Error Desc | FPT Comment
            ws.cell(row=row, column=round1_start + 1, value=test_data)   # Test Data
            ws.cell(row=row, column=round1_start + 3, value="Todo")      # Test Results
        else:
            # Round 1 sub‑columns: Tester | Test Data | Call Id | Results(Bot) | Results(data) | Error Type | Error Desc | FPT Comment
            ws.cell(row=row, column=round1_start + 1, value=test_data)   # Test Data
            ws.cell(row=row, column=round1_start + 3, value="Todo")      # Test Results (Bot responses)
            ws.cell(row=row, column=round1_start + 4, value="Todo")      # Test Results (data collection)

        # Apply group fill: column F (Expected Bot Responses) only for multi_responses; columns A-F for others
        if fill is not None:
            if is_multi_responses:
                ws.cell(row=row, column=6).fill = fill
            else:
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = fill

        return row + 1

    # ---------- Write test cases ----------
    section_header_rows: list[int] = []

    if not group_by_step:
        row = 4   # data starts at row 4: row1=Ngày test, row2=main headers, row3=sub-headers
        for idx, case in enumerate(cases, start=1):
            row = _write_data_row(row, case, idx, fill=None)
    else:
        grouped_cases: dict[str, list[dict[str, Any]]] = {}
        for case in cases:
            group_step = _extract_group_step(str(case.get("path", ""))) or "UNKNOWN"
            grouped_cases.setdefault(group_step, []).append(case)

        ordered_groups = sorted(grouped_cases.keys(), key=_step_sort_key)
        current_group_key: tuple[Any, ...] | None = None
        group_index = -1
        display_tc_index = 1
        row = 4

        for step_no in ordered_groups:
            step_name = step_name_map.get(step_no, step_no)
            section_label = f"{step_no} - {step_name}"

            # Section header row – merge across all columns (1 to total_columns)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
            cell_sec = ws.cell(row=row, column=1, value=section_label)
            cell_sec.font = SECTION_FONT
            cell_sec.fill = SECTION_FILL
            section_header_rows.append(row)
            row += 1

            for case in grouped_cases[step_no]:
                conditions = str(case.get("conditions", ""))
                steps: list[str] = case.get("steps", [])
                tc_path = str(case.get("path", ""))
                expected_action = str(case.get("expected_action_code", "N/A"))

                if use_group_fills:
                    group_key = (conditions, tuple(steps), tc_path, expected_action)
                    if group_key != current_group_key:
                        group_index += 1
                        current_group_key = group_key
                    fill = GROUP_FILLS[group_index % len(GROUP_FILLS)]
                else:
                    fill = None

                row = _write_data_row(row, case, display_tc_index, fill=fill)
                display_tc_index += 1

    # ---------- Apply formatting ----------
    _apply_borders(ws)
    _set_column_widths(ws, dynamic_keys, is_multi_responses=is_multi_responses)
    _apply_alignments(ws, dynamic_keys, section_header_rows=section_header_rows, is_multi_responses=is_multi_responses)
    _auto_fit_row_heights(ws)
    _add_data_validations(ws, dynamic_keys, ws.max_row, is_multi_responses=is_multi_responses)
    _add_conditional_formatting(ws, dynamic_keys, ws.max_row, is_multi_responses=is_multi_responses)
    if _standalone:
        _add_data_input_sheet(wb, schema_keys=schema_keys, schema_value_rows=schema_value_rows)
        _add_data_schema_sheet(wb, schema_all_rows=schema_all_rows)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert JSON test cases into an Excel file (test_cases.xlsx)."
    )
    parser.add_argument(
        "--in",
        dest="in_path",
        type=Path,
        required=True,
        help="Input JSON file containing a list of test cases.",
    )
    parser.add_argument(
        "--out",
        dest="out_path",
        type=Path,
        default=Path("output/test_cases.xlsx"),
        help="Output .xlsx path (default: output/test_cases.xlsx).",
    )
    parser.add_argument(
        "--rows",
        dest="rows_path",
        type=Path,
        default=None,
        help="Optional rows JSON (with step_no/step_name) for exact section titles.",
    )
    parser.add_argument(
        "--group-by-step",
        action="store_true",
        default=False,
        help="Group test cases by step and add section headers.",
    )
    parser.add_argument(
        "--use-group-fills",
        action="store_true",
        default=False,
        help="Apply alternating background colours to groups (requires --group-by-step).",
    )
    parser.add_argument(
        "--allow-highlight-last",
        action="store_true",
        default=False,
        help="Highlight last step in red (used by multi_responses pipeline).",
    )
    args = parser.parse_args()

    cases = _load_testcases(args.in_path)
    step_name_map = _load_step_name_map(args.in_path, rows_path=args.rows_path)
    source_rows = _load_source_rows(args.rows_path)
    export_to_excel(
        cases,
        args.out_path,
        step_name_map=step_name_map,
        group_by_step=args.group_by_step,
        use_group_fills=args.use_group_fills,
        allow_highlight_last=args.allow_highlight_last,
        source_rows=source_rows,
    )
    print(f"Wrote {len(cases)} test cases to Excel: {args.out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
