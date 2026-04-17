"""
Generate realistic customer conversation utterances for each test case using Llama 3.1 (Ollama),
and fill the Test Data column in the output Excel.
Intents in a step are separated by " \ " (backslash). Bot responses are also stored as a list
(one per step) and written to a separate column.
"""
from __future__ import annotations
from collections.abc import Iterable

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:
    from ollama import chat as ollama_chat
except ImportError:
    ollama_chat = None

# Default model for Vietnamese conversation generation
DEFAULT_MODEL = "llama3.1"
MAX_WORDS_PER_RESPONSE = 12


def _build_prompt(steps: list[str], bot_responses: list[str], tc_id: str) -> str:
    """
    Build the prompt for Llama.
    Steps may contain multiple intents separated by " \ ".
    Bot responses are shown as additional context.
    """
    steps_bullet = []
    for i, step in enumerate(steps, start=1):
        line = f"{i}. {step}"
        if i <= len(bot_responses) and bot_responses[i-1]:
            line += f"  (Bot sẽ đáp: \"{bot_responses[i-1]}\")"
        steps_bullet.append(line)
    steps_text = "\n".join(steps_bullet)

    return f"""Bạn là KHÁCH HÀNG (KH) đang nói chuyện TRỰC TIẾP qua điện thoại với tổng đài viên về KHOẢN VAY.

NHIỆM VỤ: chỉ tạo lời thoại KH (KHÔNG mô tả, KHÔNG giải thích, KHÔNG kể chuyện).

TC_ID: {tc_id}

Kịch bản (mỗi bước có thể chứa nhiều intent cách nhau bằng " \ " – hãy lấy INTENT ĐẦU TIÊN để viết):
{steps_text}

YÊU CẦU BẮT BUỘC:
- Mỗi bước = 1 câu thoại KH tương ứng đúng thứ tự
- Nếu 1 dòng có nhiều intent (" \ ") → CHỌN 1 INTENT ĐẦU TIÊN ĐỂ VIẾT
- Tổng số câu = {len(steps)}
- CHỈ là lời KH nói, không phải mô tả tình huống
- KHÔNG được nhắc lại hoặc diễn giải kịch bản
- KHÔNG được viết kiểu: "khách hàng im lặng", "người thân nói", "cuộc gọi..."
- Nếu tình huống là "im lặng" → viết câu tự nhiên thể hiện không nghe rõ (vd: alo em không nghe rõ)
- Nếu "người thân nghe hộ" → vẫn phải nói như người đang nghe điện thoại (vd: chị là mẹ chau đang nghe)

FORMAT CỨNG:
- Chỉ output danh sách đánh số (1., 2., 3...)
- Mỗi dòng 1 câu duy nhất
- KHÔNG thêm tiêu đề, KHÔNG markdown, KHÔNG giải thích
- KHÔNG ký tự đặc biệt: *^$#@%&?!
- KHÔNG dấu chấm, chấm hỏi, chấm than cuối câu
- KHÔNG dấu phẩy
- viết thường toàn bộ
- tối đa {MAX_WORDS_PER_RESPONSE} từ mỗi câu

VÍ DỤ:

Input:
- KH im lặng lần 1
- KH im lặng lần 2

Output:
1. alo em nghe không rõ anh nói lại giúp em với
2. dạ vẫn chưa nghe rõ anh nói lớn hơn chút được không

Đầu ra:"""


def _parse_numbered_response(raw: str, expected_count: int) -> list[str]:
    """Extract numbered lines (1. ... 2. ...) from model output; return list of response texts."""
    lines: list[str] = []
    # Match lines like "1. text" or "1) text"
    pattern = re.compile(r"^\s*(\d+)[.)]\s*(.+)$", re.MULTILINE)
    for m in pattern.finditer(raw):
        lines.append(m.group(2).strip())
    if len(lines) >= expected_count:
        return lines[:expected_count]
    # Fallback: split by newline and strip leading numbers
    if not lines:
        for line in raw.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            # Remove leading "1. " or "1) "
            cleaned = re.sub(r"^\s*\d+[.)]\s*", "", line).strip()
            if cleaned:
                lines.append(cleaned)
    return lines[:expected_count] if lines else [raw.strip()]


def _generate_with_ollama(steps: list[str], bot_responses: list[str], tc_id: str, model: str) -> str:
    """Call Ollama (Llama 3.1) to generate one conversation; return numbered lines as single string."""
    if not steps:
        return ""
    if ollama_chat is None:
        raise RuntimeError(
            "Ollama Python package not installed. Run: pip install ollama"
        )
    prompt = _build_prompt(steps, bot_responses, tc_id)
    response = ollama_chat(
        model=model,
        messages=[{"role": "user", "content": prompt}],
    )
    content = (response.get("message") or {}).get("content") or ""
    parsed = _parse_numbered_response(content, len(steps))
    # Ensure we have exactly expected_count lines; pad or truncate
    while len(parsed) < len(steps):
        parsed.append("(không sinh được)")
    result = "\n".join(f"{i+1}. {p}" for i, p in enumerate(parsed[: len(steps)]))
    return result


def _auto_fit_column(ws: Worksheet, col_idx: int, extra: int = 2) -> None:
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            for part in str(cell.value).split("\n"):
                max_len = max(max_len, len(part))
    ws.column_dimensions[col_letter].width = max_len + extra


def fill_excel_from_json(
    json_path: Path,
    excel_path: Path,
    model: str = DEFAULT_MODEL,
    limit: int | None = None,
) -> None:
    """Load TC JSON, generate test data for each, write Excel with Test Data and Bot Responses columns."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of test cases.")
    cases: list[dict[str, Any]] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            continue
        tc_id = str(item.get("tc_id", f"TC_{i+1:03d}"))
        steps = item.get("steps", [])
        if not isinstance(steps, Iterable):
            steps = []
        bot_responses = item.get("bot_responses", [])
        if not isinstance(bot_responses, Iterable):
            bot_responses = []
        clean_steps = [
            str(s).strip()
            for s in steps
            if str(s).strip() and not str(s).strip().startswith("(cycle to")
        ]
        clean_bot_responses = [str(r).strip() for r in bot_responses]
        # Ensure lengths match
        while len(clean_bot_responses) < len(clean_steps):
            clean_bot_responses.append("")
        expected_action_code = str(item.get("expected_action_code", ""))
        cases.append({
            "tc_id": tc_id,
            "steps": clean_steps,
            "bot_responses": clean_bot_responses,
            "expected_action_code": expected_action_code,
            "path": str(item.get("path", "")),
        })
    if limit is not None:
        cases = cases[:limit]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestCases"
    ws["A1"] = "TC_ID"
    ws["B1"] = "Test Scenario"
    ws["C1"] = "Bot Responses"
    ws["D1"] = "Path"
    ws["E1"] = "Expected Action Code"
    ws["F1"] = "Test Data"

    for row_idx, case in enumerate(cases, start=2):
        tc_id = case["tc_id"]
        steps = case["steps"]
        bot_responses = case["bot_responses"]
        scenario = "\n".join(f"{i}. {s}" for i, s in enumerate(steps, start=1))
        # Format bot responses as numbered list (each cell may contain backslash-separated variants)
        bot_responses_text = "\n".join(
            f"{i}. {r}" for i, r in enumerate(bot_responses, start=1) if r
        )
        test_data = _generate_with_ollama(steps, bot_responses, tc_id, model)
        ws.cell(row=row_idx, column=1, value=tc_id)
        cell_b = ws.cell(row=row_idx, column=2, value=scenario)
        cell_b.alignment = Alignment(wrap_text=True)
        cell_c = ws.cell(row=row_idx, column=3, value=bot_responses_text)
        cell_c.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=case.get("path", ""))
        ws.cell(row=row_idx, column=5, value=case["expected_action_code"])
        cell_f = ws.cell(row=row_idx, column=6, value=test_data)
        cell_f.alignment = Alignment(wrap_text=True)
        print(f"  Generated: {tc_id} ({len(steps)} steps)")

    for c in range(1, 7):
        _auto_fit_column(ws, c, extra=4)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def fill_excel_from_excel(
    excel_path: Path,
    out_path: Path,
    model: str = DEFAULT_MODEL,
    limit: int | None = None,
) -> None:
    """Read existing Excel, fill Test Data column using Llama, and add Bot Responses column if missing."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    max_row = ws.max_row
    if limit is not None:
        max_row = min(max_row, 1 + limit)

    # Find columns by header names
    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val is None:
            continue
        header_to_col[str(header_val).strip()] = col

    tc_id_col = header_to_col.get("TC_ID", 1)
    scenario_col = header_to_col.get("Test Scenario", 2)
    bot_responses_col = header_to_col.get("Bot Responses", 0)  # 0 if not present
    path_col = header_to_col.get("Path", 3)
    action_col = header_to_col.get("Expected Action Code", 4)
    test_data_col = header_to_col.get("Test Data", 5)

    # If Bot Responses column is missing, we need to add it and adjust columns
    if bot_responses_col == 0:
        # Insert a new column after Test Scenario
        ws.insert_cols(scenario_col + 1)
        ws.cell(row=1, column=scenario_col + 1, value="Bot Responses")
        bot_responses_col = scenario_col + 1
        # Update subsequent column indices
        if path_col >= bot_responses_col:
            path_col += 1
        if action_col >= bot_responses_col:
            action_col += 1
        if test_data_col >= bot_responses_col:
            test_data_col += 1

    # Collect data for each row
    rows_data = []
    for row in range(2, max_row + 1):
        tc_id = ws.cell(row=row, column=tc_id_col).value or ""
        scenario_cell = ws.cell(row=row, column=scenario_col).value or ""
        bot_responses_cell = ws.cell(row=row, column=bot_responses_col).value or ""
        # Parse steps from scenario (numbered lines)
        steps: list[str] = []
        for line in str(scenario_cell).strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r"^\s*\d+[.)]\s*", "", line).strip()
            if cleaned:
                steps.append(cleaned)
        # Parse bot responses (numbered lines)
        bot_responses: list[str] = []
        for line in str(bot_responses_cell).strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r"^\s*\d+[.)]\s*", "", line).strip()
            if cleaned:
                bot_responses.append(cleaned)
        # Ensure lengths match
        while len(bot_responses) < len(steps):
            bot_responses.append("")
        if not steps:
            continue
        rows_data.append({
            "tc_id": tc_id,
            "steps": steps,
            "bot_responses": bot_responses,
            "path": ws.cell(row=row, column=path_col).value or "",
            "expected_action": ws.cell(row=row, column=action_col).value or "",
        })

    # Build new workbook with consistent columns
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "TestCases"
    new_ws["A1"] = "TC_ID"
    new_ws["B1"] = "Test Scenario"
    new_ws["C1"] = "Bot Responses"
    new_ws["D1"] = "Path"
    new_ws["E1"] = "Expected Action Code"
    new_ws["F1"] = "Test Data"

    for row_idx, case in enumerate(rows_data, start=2):
        tc_id = case["tc_id"]
        steps = case["steps"]
        bot_responses = case["bot_responses"]
        scenario = "\n".join(f"{i}. {s}" for i, s in enumerate(steps, start=1))
        bot_responses_text = "\n".join(
            f"{i}. {r}" for i, r in enumerate(bot_responses, start=1) if r
        )
        test_data = _generate_with_ollama(steps, bot_responses, str(tc_id), model)
        new_ws.cell(row=row_idx, column=1, value=tc_id)
        cell_b = new_ws.cell(row=row_idx, column=2, value=scenario)
        cell_b.alignment = Alignment(wrap_text=True)
        cell_c = new_ws.cell(row=row_idx, column=3, value=bot_responses_text)
        cell_c.alignment = Alignment(wrap_text=True)
        new_ws.cell(row=row_idx, column=4, value=case["path"])
        new_ws.cell(row=row_idx, column=5, value=case["expected_action"])
        cell_f = new_ws.cell(row=row_idx, column=6, value=test_data)
        cell_f.alignment = Alignment(wrap_text=True)
        print(f"  Generated: {tc_id} ({len(steps)} steps)")

    for c in range(1, 7):
        _auto_fit_column(new_ws, c, extra=4)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    new_wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate test data (customer utterances) with Llama 3.1 and fill Excel Test Data column."
    )
    parser.add_argument(
        "--in",
        dest="in_path",
        type=Path,
        required=True,
        help="Input: JSON (tcgen output) or Excel (tc_to_excel output).",
    )
    parser.add_argument(
        "--out",
        dest="out_path",
        type=Path,
        default=None,
        help="Output Excel path. If input is Excel, defaults to overwriting input.",
    )
    parser.add_argument(
        "--model",
        type=str,
        default=DEFAULT_MODEL,
        help=f"Ollama model name (default: {DEFAULT_MODEL}).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N test cases (for testing).",
    )
    args = parser.parse_args()

    in_path = args.in_path
    if not in_path.exists():
        print(f"Error: Input file not found: {in_path}")
        return 1
    if ollama_chat is None:
        print("Error: Install Ollama Python client: pip install ollama")
        return 1

    out_path = args.out_path
    if out_path is None:
        out_path = in_path if in_path.suffix.lower() == ".xlsx" else Path("output/test_cases.xlsx")

    if in_path.suffix.lower() == ".json":
        print(f"Reading JSON: {in_path}")
        print(f"Writing Excel with test data: {out_path}")
        fill_excel_from_json(in_path, out_path, model=args.model, limit=args.limit)
    else:
        print(f"Reading Excel: {in_path}")
        print("Filling Test Data column and adding Bot Responses if needed...")
        fill_excel_from_excel(in_path, out_path, model=args.model, limit=args.limit)

    print(f"Done. Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())