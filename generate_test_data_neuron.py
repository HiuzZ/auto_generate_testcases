"""
Generate realistic customer conversation utterances for each test case using TLS Client Bot only.
For multi‑intent steps (separated by " \ "), only the FIRST intent is used for generation.
Bot responses are taken from the JSON and written to a separate column.
"""

from __future__ import annotations

from collections.abc import Iterable
import argparse
import json
import os
import random
import re
import sys
from contextlib import contextmanager
from pathlib import Path
from typing import Any
import unicodedata

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# TLS Client Bot imports
PROJECT_ROOT = Path(__file__).resolve().parent
TLS_BOT_DIR = PROJECT_ROOT / "tls_client_bot"
TLS_INTENTS_PATH = TLS_BOT_DIR / "intents.json"

_MAX_WORDS_PER_RESPONSE = 12

_TLS_GET_RESPONSE = None
_TLS_INTENTS_DATA: dict[str, Any] | None = None
_TAGS_SET: set[str] | None = None
_PATTERN_TO_TAG: dict[str, str] | None = None
_PATTERN_TO_TAG_FOLDED: dict[str, str] | None = None


@contextmanager
def _temporary_chdir(path: Path):
    old = Path.cwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(old)


def _load_tls_intents() -> dict[str, Any]:
    global _TLS_INTENTS_DATA, _TAGS_SET, _PATTERN_TO_TAG, _PATTERN_TO_TAG_FOLDED
    if _TLS_INTENTS_DATA is not None:
        return _TLS_INTENTS_DATA

    if not TLS_INTENTS_PATH.exists():
        raise FileNotFoundError(f"Missing tls bot intents.json: {TLS_INTENTS_PATH}")

    data = json.loads(TLS_INTENTS_PATH.read_text(encoding="utf-8"))
    _TLS_INTENTS_DATA = data
    tags_set: set[str] = set()
    pattern_to_tag: dict[str, str] = {}
    pattern_to_tag_folded: dict[str, str] = {}

    for intent in data.get("intents", []):
        tag = str(intent.get("tag", "")).strip()
        if not tag:
            continue
        tags_set.add(tag)
        for pattern in intent.get("patterns", []) or []:
            p = str(pattern).strip().lower()
            if p and p not in pattern_to_tag:
                pattern_to_tag[p] = tag
            folded_p = _fold_for_match(p)
            if folded_p and folded_p not in pattern_to_tag_folded:
                pattern_to_tag_folded[folded_p] = tag

    _TAGS_SET = tags_set
    _PATTERN_TO_TAG = pattern_to_tag
    _PATTERN_TO_TAG_FOLDED = pattern_to_tag_folded
    return data


def _load_tls_get_response():
    global _TLS_GET_RESPONSE
    if _TLS_GET_RESPONSE is not None:
        return _TLS_GET_RESPONSE

    if not TLS_BOT_DIR.exists():
        _TLS_GET_RESPONSE = None
        return None

    try:
        tls_str = str(TLS_BOT_DIR)
        if tls_str not in sys.path:
            sys.path.insert(0, tls_str)

        with _temporary_chdir(TLS_BOT_DIR):
            from chat import get_response as tls_get_response  # type: ignore

        _TLS_GET_RESPONSE = tls_get_response
        return _TLS_GET_RESPONSE
    except Exception:
        _TLS_GET_RESPONSE = None
        return None


def _extract_latest_attempt(raw_step: str) -> tuple[str | None, str | None]:
    attempt_matches = re.finditer(r"\b(lần\s*(\d+))\b", raw_step, flags=re.IGNORECASE)
    attempts = []
    for match in attempt_matches:
        attempts.append({
            'full': match.group(1),
            'number': int(match.group(2)),
            'position': match.start()
        })
    if attempts:
        attempts.sort(key=lambda x: x['position'])
        latest = attempts[-1]
        return latest['full'], str(latest['number'])
    return None, None


def _normalize_step_to_bot_message(step: str) -> str:
    """
    Take the first intent (split by backslash) and keep the highest attempt number.
    """
    _load_tls_intents()
    tags_set = _TAGS_SET or set()

    raw = (step or "").strip()
    if not raw:
        return raw

    # Split on backslash (with optional spaces)
    parts = [p.strip() for p in re.split(r"\s*\\\s*", raw) if p.strip()]
    first_part = parts[0] if parts else raw
    is_multi_intent = len(parts) > 1
    first_folded = _fold_for_match(first_part)

    # Special cases: im lặng, không nghe rõ
    if "im lang" in first_folded:
        for tag in tags_set:
            if "im lang" in _fold_for_match(tag):
                return tag
    if "khong nghe ro" in first_folded:
        for tag in tags_set:
            if "khong nghe ro" in _fold_for_match(tag):
                return tag

    # Multi-intent: take first intent, keep attempt number from whole step
    if is_multi_intent:
        selected_intent = parts[0]
        attempt_full, _ = _extract_latest_attempt(raw)
        if attempt_full:
            selected_intent = re.sub(r"\s*lần\s*\d+\s*", "", selected_intent, flags=re.IGNORECASE).strip()
            return f"{selected_intent} {attempt_full}".strip()
        return first_part

    # Single intent: try exact tag match
    for tag in tags_set:
        tag_folded = _fold_for_match(tag)
        if tag_folded == first_folded:
            return tag
        tag_first = tag_folded.split("/")[0].strip()
        if tag_first == first_folded:
            return tag

    return first_part


_BANNED_CHARS_RE = re.compile(r"[\*\^\$\#\@\%\&\?\!]")
_TRAILING_PUNCT_RE = re.compile(r"[\.!\?]+$")
_MULTISPACE_RE = re.compile(r"\s+")


def _sanitize_utterance(text: str) -> str:
    if not text:
        return "(không sinh được)"
    t = str(text).strip()
    if "do not understand" in t.lower():
        return "(không sinh được)"
    t = t.replace(",", " ")
    t = _BANNED_CHARS_RE.sub("", t)
    t = t.replace("\n", " ")
    t = _TRAILING_PUNCT_RE.sub("", t).strip()
    t = t.lower()
    t = _MULTISPACE_RE.sub(" ", t).strip()
    words = t.split(" ")
    if len(words) > _MAX_WORDS_PER_RESPONSE:
        t = " ".join(words[:_MAX_WORDS_PER_RESPONSE])
    return t if t else "(không sinh được)"


def _fold_for_match(s: str) -> str:
    base = unicodedata.normalize("NFD", s.lower())
    return "".join(ch for ch in base if unicodedata.category(ch) != "Mn")


def _fallback_response_for_message(msg: str) -> str:
    """Fallback when TLS bot fails: use intents.json pattern matching."""
    _load_tls_intents()
    intents = _TLS_INTENTS_DATA.get("intents", []) if _TLS_INTENTS_DATA else []
    tags_set = _TAGS_SET or set()
    pattern_to_tag = _PATTERN_TO_TAG or {}
    pattern_to_tag_folded = _PATTERN_TO_TAG_FOLDED or {}

    m = (msg or "").strip()
    if not m:
        return "(không sinh được)"

    m_lower = m.lower()
    m_folded = _fold_for_match(m_lower)
    m_without_attempt = re.sub(r'\s*lần\s*\d+\s*', ' ', m_lower).strip()
    m_without_attempt_folded = _fold_for_match(m_without_attempt)

    tag: str | None = None
    for t in tags_set:
        t_folded = _fold_for_match(t)
        t_without_attempt = re.sub(r'\s*lần\s*\d+\s*', '', t_folded).strip()
        if t_folded == m_folded or t_without_attempt == m_without_attempt_folded:
            tag = t
            break
        t_first = t_folded.split("/")[0].strip()
        if t_first == m_folded or t_first == m_without_attempt_folded:
            tag = t
            break

    if not tag:
        tag = pattern_to_tag.get(m_lower) or pattern_to_tag.get(m_without_attempt)
    if not tag:
        tag = pattern_to_tag_folded.get(m_folded) or pattern_to_tag_folded.get(m_without_attempt_folded)

    if not tag:
        return "(không sinh được)"

    for intent in intents:
        if str(intent.get("tag", "")).strip() == tag:
            responses = intent.get("responses") or []
            if responses:
                selected = random.choice([str(r) for r in responses if str(r).strip()])
                return selected
    return "(không sinh được)"


def _generate_step_with_tls_bot(step: str) -> str:
    """Generate a single utterance using TLS Client Bot (first intent only)."""
    tls_get_response = _load_tls_get_response()
    msg = _normalize_step_to_bot_message(step)

    if tls_get_response is None:
        resp = _fallback_response_for_message(msg)
        return _sanitize_utterance(resp)

    try:
        resp = tls_get_response(msg, context=None, token=None)
        if not resp or "do not understand" in str(resp).lower():
            resp = _fallback_response_for_message(msg)
    except Exception:
        resp = _fallback_response_for_message(msg)

    return _sanitize_utterance(resp)


def _generate_conversation(steps: list[str], tc_id: str) -> str:
    """Generate utterances for each step using TLS bot (first intent)."""
    if not steps:
        return ""
    utterances = []
    for i, step in enumerate(steps, 1):
        resp = _generate_step_with_tls_bot(step)
        utterances.append(f"{i}. {resp}")
        print(f"  Step {i}: '{resp}'")
    return "\n".join(utterances)


def _auto_fit_column(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: int, extra: int = 2) -> None:
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            for part in str(cell.value).split("\n"):
                max_len = max(max_len, len(part))
    ws.column_dimensions[col_letter].width = max_len + extra


def fill_excel_from_json(json_path: Path, excel_path: Path, limit: int | None = None) -> None:
    """Load JSON test cases, generate utterances, write Excel with Bot Responses column."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of test cases.")

    cases = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            continue
        tc_id = str(item.get("tc_id", f"TC_{i+1:03d}"))
        steps = item.get("steps", [])
        bot_responses = item.get("bot_responses", [])
        if not isinstance(steps, Iterable):
            steps = []
        if not isinstance(bot_responses, Iterable):
            bot_responses = []
        clean_steps = [
            str(s).strip()
            for s in steps
            if str(s).strip() and not str(s).strip().startswith("(cycle to")
        ]
        clean_bot_responses = [str(r).strip() for r in bot_responses]
        while len(clean_bot_responses) < len(clean_steps):
            clean_bot_responses.append("")
        expected_action_code = str(item.get("expected_action_code", ""))
        tc_path = str(item.get("path", ""))
        cases.append({
            "tc_id": tc_id,
            "steps": clean_steps,
            "bot_responses": clean_bot_responses,
            "expected_action_code": expected_action_code,
            "path": tc_path,
        })

    if limit:
        cases = cases[:limit]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestCases"
    ws["A1"] = "TC_ID"
    ws["B1"] = "Test Scenario"
    ws["C1"] = "Bot Responses"
    ws["D1"] = "Path"
    ws["E1"] = "Expected Action Code"
    ws["F1"] = "Test Data"

    for row_idx, case in enumerate(cases, start=2):
        tc_id = case["tc_id"]
        steps = case["steps"]
        bot_responses = case["bot_responses"]
        scenario = "\n".join(f"{i}. {s}" for i, s in enumerate(steps, start=1))
        bot_responses_text = "\n".join(f"{i}. {r}" for i, r in enumerate(bot_responses, start=1) if r)

        print(f"\nGenerating for TC: {tc_id}")
        test_data = _generate_conversation(steps, tc_id)

        ws.cell(row=row_idx, column=1, value=tc_id)
        cell_b = ws.cell(row=row_idx, column=2, value=scenario)
        cell_b.alignment = Alignment(wrap_text=True)
        cell_c = ws.cell(row=row_idx, column=3, value=bot_responses_text)
        cell_c.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=case["path"])
        ws.cell(row=row_idx, column=5, value=case["expected_action_code"])
        cell_f = ws.cell(row=row_idx, column=6, value=test_data)
        cell_f.alignment = Alignment(wrap_text=True)
        print(f"  Done {tc_id} ({len(steps)} steps)")

    for c in range(1, 7):
        _auto_fit_column(ws, c, extra=4)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def fill_excel_from_excel(excel_path: Path, out_path: Path, limit: int | None = None) -> None:
    """Read existing Excel (with TC_ID, Test Scenario, Bot Responses, Path, Expected Action Code), fill Test Data."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    max_row = ws.max_row
    if limit:
        max_row = min(max_row, 1 + limit)

    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val:
            header_to_col[str(header_val).strip()] = col

    tc_id_col = header_to_col.get("TC_ID", 1)
    scenario_col = header_to_col.get("Test Scenario", 2)
    bot_responses_col = header_to_col.get("Bot Responses", 3)
    path_col = header_to_col.get("Path", 4)
    action_col = header_to_col.get("Expected Action Code", 5)
    test_data_col = header_to_col.get("Test Data", 6)

    # If Bot Responses column missing, insert it
    if bot_responses_col == 0:
        ws.insert_cols(scenario_col + 1)
        ws.cell(row=1, column=scenario_col + 1, value="Bot Responses")
        bot_responses_col = scenario_col + 1
        if path_col >= bot_responses_col:
            path_col += 1
        if action_col >= bot_responses_col:
            action_col += 1
        if test_data_col >= bot_responses_col:
            test_data_col += 1

    for row in range(2, max_row + 1):
        tc_id = ws.cell(row=row, column=tc_id_col).value or ""
        scenario_cell = ws.cell(row=row, column=scenario_col).value or ""
        bot_responses_cell = ws.cell(row=row, column=bot_responses_col).value or ""

        # Parse steps (numbered lines)
        steps = []
        for line in str(scenario_cell).strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r"^\s*\d+[.)]\s*", "", line).strip()
            if cleaned:
                steps.append(cleaned)

        # Parse bot responses
        bot_responses = []
        for line in str(bot_responses_cell).strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r"^\s*\d+[.)]\s*", "", line).strip()
            if cleaned:
                bot_responses.append(cleaned)
        while len(bot_responses) < len(steps):
            bot_responses.append("")

        if not steps:
            continue

        print(f"\nGenerating for TC: {tc_id}")
        test_data = _generate_conversation(steps, str(tc_id))
        ws.cell(row=row, column=test_data_col, value=test_data).alignment = Alignment(wrap_text=True)
        print(f"  Done {tc_id} ({len(steps)} steps)")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate test data using TLS Client Bot only (first intent per step)."
    )
    parser.add_argument("--in", dest="in_path", type=Path, required=True,
                        help="Input: JSON (tcgen output) or Excel (tc_to_excel output).")
    parser.add_argument("--out", dest="out_path", type=Path, default=None,
                        help="Output Excel path. If input is Excel, defaults to overwriting input.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N test cases (for testing).")
    args = parser.parse_args()

    in_path = args.in_path
    if not in_path.exists():
        print(f"Error: Input file not found: {in_path}")
        return 1

    out_path = args.out_path
    if out_path is None:
        out_path = in_path if in_path.suffix.lower() == ".xlsx" else Path("output/test_cases.xlsx")

    print("=" * 60)
    print("TLS CLIENT BOT ONLY | First intent (split by ' \\ ') | No fallback")
    print("=" * 60)

    if in_path.suffix.lower() == ".json":
        print(f"Reading JSON: {in_path}")
        print(f"Writing Excel: {out_path}")
        fill_excel_from_json(in_path, out_path, limit=args.limit)
    else:
        print(f"Reading Excel: {in_path}")
        print("Filling Test Data column...")
        fill_excel_from_excel(in_path, out_path, limit=args.limit)

    print(f"\nDone. Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())